from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import threading
import traceback
import unicodedata
from collections import defaultdict, deque
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz
import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox

try:
    import ttkbootstrap as ttk
    from ttkbootstrap.dialogs import Messagebox as BootMessagebox

    HAS_TTKBOOTSTRAP = True
except Exception:
    from tkinter import ttk

    HAS_TTKBOOTSTRAP = False
    BootMessagebox = None


ProgressCallback = Callable[[str, float | None], None]
APP_VERSION = "v2.1"
REQUIRED_PYTHON_MAJOR = 3
REQUIRED_PYTHON_MINOR = 12
UI_STATE_FILE = Path.home() / ".matcher_matcher_ui_state.json"
QUICK_PRESETS = {
    "Equilibrado": {
        "accept_score": "92",
        "review_score": "85",
        "min_gap_for_accept": "4",
        "top_candidates_to_keep": "5",
        "allow_reuse_t2_matches": False,
        "max_matches_per_t2_name": "3",
    },
    "Conservador": {
        "accept_score": "95",
        "review_score": "90",
        "min_gap_for_accept": "6",
        "top_candidates_to_keep": "4",
        "allow_reuse_t2_matches": False,
        "max_matches_per_t2_name": "2",
    },
    "Alta Cobertura": {
        "accept_score": "90",
        "review_score": "82",
        "min_gap_for_accept": "3",
        "top_candidates_to_keep": "7",
        "allow_reuse_t2_matches": True,
        "max_matches_per_t2_name": "4",
    },
}
UI_COLORS = {
    "bg": "#1E2A30",
    "panel": "#263842",
    "panel_alt": "#304751",
    "card": "#354F5A",
    "field": "#3E5B67",
    "accent": "#2F80ED",
    "accent_alt": "#20BF6B",
    "text": "#E6F1F4",
    "muted": "#A6BEC4",
    "border": "#4E6873",
    "btn_primary": "#2F80ED",
    "btn_success": "#20BF6B",
    "btn_warning": "#F39C12",
    "btn_danger": "#E74C3C",
    "btn_info": "#00A8CC",
}
DEFAULT_MATCH_COLORS = {
    "color_exact": "FF4CAF50",
    "color_match": "FF2F80ED",
    "color_review": "FFF39C12",
    "color_no_match": "FFE57373",
    "color_excess_left": "FF64B5F6",
}
COLOR_NAME_TO_FILL = {
    "Verde (Exato)": "FF4CAF50",
    "Azul (Match aceito)": "FF2F80ED",
    "Laranja (Revisão)": "FFF39C12",
    "Vermelho (Sem match)": "FFE57373",
    "Azul claro (Excedente Excel 1)": "FF64B5F6",
}
DEFAULT_MATCH_COLOR_LABELS = {
    "color_exact": "Verde (Exato)",
    "color_match": "Azul (Match aceito)",
    "color_review": "Laranja (Revisão)",
    "color_no_match": "Vermelho (Sem match)",
    "color_excess_left": "Azul claro (Excedente Excel 1)",
}
MATCH_COLOR_KEYS = tuple(DEFAULT_MATCH_COLORS.keys())
MATCH_COLOR_LABEL_OPTIONS = list(COLOR_NAME_TO_FILL.keys())
COLOR_FILL_TO_LABEL = {fill: label for label, fill in COLOR_NAME_TO_FILL.items()}
DEFAULT_SCORE_WEIGHTS = {
    "weight_token_set": 27.0,
    "weight_partial": 21.0,
    "weight_sort": 15.0,
    "weight_prefix": 15.0,
    "weight_ordered_chars": 14.0,
    "weight_aligned_chars": 8.0,
}


# =========================
# MATCHING CORE
# =========================


@dataclass
class AnalysisResult:
    config: dict[str, Any]
    results_df: pd.DataFrame
    candidates_df: pd.DataFrame
    catalog_df: pd.DataFrame
    quota_df: pd.DataFrame
    summary_df: pd.DataFrame
    review_df: pd.DataFrame
    preview_df: pd.DataFrame
    source_df: pd.DataFrame
    target_df: pd.DataFrame
    grouped_reconciliation_df: pd.DataFrame


class FlowEdge:
    def __init__(self, to_node: int, rev_index: int, capacity: int, cost: int) -> None:
        self.to_node = to_node
        self.rev_index = rev_index
        self.capacity = capacity
        self.cost = cost


def emit_progress(callback: ProgressCallback | None, message: str, percent: float | None = None) -> None:
    if callback:
        callback(message, percent)


def ensure_supported_python_version() -> None:
    current = (sys.version_info.major, sys.version_info.minor)
    required = (REQUIRED_PYTHON_MAJOR, REQUIRED_PYTHON_MINOR)
    if current != required:
        current_text = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
        required_text = f"{REQUIRED_PYTHON_MAJOR}.{REQUIRED_PYTHON_MINOR}.x"
        raise RuntimeError(
            f"Versão de Python incompatível: {current_text}. Este programa exige Python {required_text}."
        )


def excel_col_to_index(col: str) -> int:
    col = str(col).strip().upper()
    if not col:
        raise ValueError("A letra da coluna não pode ficar vazia.")
    value = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Coluna do Excel inválida: {col}")
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return value - 1


def normalize_name(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def first_token(text: str) -> str:
    parts = text.split()
    return parts[0] if parts else ""


def last_token(text: str) -> str:
    parts = text.split()
    return parts[-1] if parts else ""


def token_set(text: str) -> set[str]:
    return set(text.split()) if text else set()


def safe_float(value: Any) -> float | None:
    if pd.isna(value) or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def parse_bool(value: Any, field_name: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value in {0, 0.0}:
            return False
        if value in {1, 1.0}:
            return True
    text = str(value or "").strip().lower()
    if text in {"1", "true", "t", "yes", "y", "sim", "s", "on"}:
        return True
    if text in {"0", "false", "f", "no", "n", "nao", "não", "off", ""}:
        return False
    raise ValueError(f"O campo '{field_name}' deve ser booleano (true/false). Valor recebido: {value!r}")


def add_flag(flags: list[str], flag: str, enabled: bool) -> None:
    if enabled and flag not in flags:
        flags.append(flag)


def flags_to_text(flags: list[str]) -> str:
    return "; ".join(sorted(set(flag for flag in flags if flag)))


def resolve_score_weights(config: dict[str, Any] | None = None) -> dict[str, float]:
    if config is None:
        return dict(DEFAULT_SCORE_WEIGHTS)

    weights = {}
    for key, default_value in DEFAULT_SCORE_WEIGHTS.items():
        raw_value = config.get(key, default_value)
        try:
            weights[key] = float(raw_value)
        except Exception:
            weights[key] = float(default_value)

    positive_total = sum(max(0.0, value) for value in weights.values())
    if positive_total <= 0:
        return dict(DEFAULT_SCORE_WEIGHTS)
    return {key: max(0.0, value) / positive_total for key, value in weights.items()}


def aligned_character_ratio(left: str, right: str) -> float:
    max_len = max(len(left), len(right))
    if max_len == 0:
        return 100.0
    same_position = sum(1 for a, b in zip(left, right) if a == b)
    return round((same_position / max_len) * 100, 2)


def ordered_character_ratio(left: str, right: str) -> float:
    if not left and not right:
        return 100.0
    return round(SequenceMatcher(None, left, right).ratio() * 100, 2)


def score_candidate(
    full_name: str,
    external_name: str,
    max_external_chars: int,
    match2_left: str = "",
    match2_right: str = "",
    match2_prefix_chars: int = 8,
    match2_weight: float = 0.2,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    full_tokens = token_set(full_name)
    ext_tokens = token_set(external_name)
    weights = resolve_score_weights(config)

    same_first = first_token(full_name) == first_token(external_name) and first_token(full_name) != ""
    same_last = last_token(full_name) == last_token(external_name) and last_token(full_name) != ""
    ext_subset_in_full = bool(ext_tokens) and ext_tokens.issubset(full_tokens)
    full_subset_in_ext = bool(full_tokens) and full_tokens.issubset(ext_tokens)
    starts_like = full_name.startswith(external_name) or external_name.startswith(full_name)
    length_gap = abs(len(full_name) - len(external_name))
    same_name_length = len(full_name) == len(external_name)

    score_token_set = float(fuzz.token_set_ratio(full_name, external_name))
    score_partial = float(fuzz.partial_ratio(full_name, external_name))
    score_sort = float(fuzz.token_sort_ratio(full_name, external_name))
    score_prefix = float(fuzz.ratio(full_name[:max_external_chars], external_name[:max_external_chars]))
    score_ordered_chars = ordered_character_ratio(full_name, external_name)
    score_aligned_chars = aligned_character_ratio(full_name, external_name)
    m2_left = str(match2_left or "").strip().upper()
    m2_right = str(match2_right or "").strip().upper()
    prefix_chars = max(int(match2_prefix_chars or 1), 1)
    m2_prefix_left = m2_left[:prefix_chars]
    m2_prefix_right = m2_right[:prefix_chars]
    match2_equal_prefix = bool(m2_prefix_left and m2_prefix_right and m2_prefix_left == m2_prefix_right)
    if m2_left or m2_right:
        match2_score = ordered_character_ratio(m2_prefix_left, m2_prefix_right)
    else:
        match2_score = 0.0

    score = (
        weights["weight_token_set"] * score_token_set
        + weights["weight_partial"] * score_partial
        + weights["weight_sort"] * score_sort
        + weights["weight_prefix"] * score_prefix
        + weights["weight_ordered_chars"] * score_ordered_chars
        + weights["weight_aligned_chars"] * score_aligned_chars
    )
    if same_first:
        score += 6
    if same_last:
        score += 4
    if ext_subset_in_full or full_subset_in_ext:
        score += 8
    if starts_like:
        score += 4
    if score_ordered_chars >= 94:
        score += 3
    if score_aligned_chars >= 88:
        score += 2
    score -= min(length_gap * float(config.get("length_gap_penalty_per_char", 0.5) if config else 0.5), float(config.get("max_length_gap_penalty", 10.0) if config else 10.0))
    if same_first and not same_last and len(full_tokens) >= 2 and len(ext_tokens) >= 2:
        score -= float(config.get("missing_surname_penalty", 3.0) if config else 3.0)

    score = min(max(score, 0.0), 100.0)
    score_composite = min(max(score + (match2_score * float(match2_weight)), 0.0), 100.0)
    structure_ok = same_first and (
        same_last
        or ext_subset_in_full
        or score_token_set >= 88
        or score_ordered_chars >= 90
        or score_aligned_chars >= 86
    )
    needs_length_review = bool(
        same_first
        and score_token_set >= 95.0
        and (ext_subset_in_full or full_subset_in_ext or starts_like)
        and (not same_name_length or score_aligned_chars < 100.0)
    )

    return {
        "score": round(score, 2),
        "score_composite": round(score_composite, 2),
        "same_first": same_first,
        "same_last": same_last,
        "ext_subset_in_full": ext_subset_in_full,
        "full_subset_in_ext": full_subset_in_ext,
        "starts_like": starts_like,
        "score_token_set": round(score_token_set, 2),
        "score_partial": round(score_partial, 2),
        "score_sort": round(score_sort, 2),
        "score_prefix": round(score_prefix, 2),
        "score_ordered_chars": round(score_ordered_chars, 2),
        "score_aligned_chars": round(score_aligned_chars, 2),
        "same_name_length": same_name_length,
        "name_length_gap": length_gap,
        "needs_length_review": needs_length_review,
        "structure_ok": structure_ok,
        "match2_left": m2_left,
        "match2_right": m2_right,
        "match2_prefix_left": m2_prefix_left,
        "match2_prefix_right": m2_prefix_right,
        "match2_equal_prefix": match2_equal_prefix,
        "match2_score": round(match2_score, 2),
    }


def build_summary(df: pd.DataFrame, status_column: str = "final_status") -> pd.DataFrame:
    total = max(len(df), 1)
    summary = (
        df[status_column]
        .fillna("SEM_STATUS")
        .value_counts(dropna=False)
        .rename_axis("status")
        .reset_index(name="quantidade")
    )
    summary["percentual"] = (summary["quantidade"] / total * 100).round(2)
    return summary


def _autosize_columns(ws, max_width: int = 56) -> None:
    for col in ws.columns:
        if not col:
            continue
        letter = col[0].column_letter
        best = 0
        for cell in col[:3000]:
            try:
                value = "" if cell.value is None else str(cell.value)
            except Exception:
                value = ""
            best = max(best, len(value))
        ws.column_dimensions[letter].width = min(max(best + 2, 10), max_width)


def _find_header_index(ws, header_name: str) -> int | None:
    target = header_name.strip()
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip() == target:
            return idx
    return None


def format_output_workbook(output_file: Path, result: AnalysisResult, results_startrow: int) -> None:
    wb = load_workbook(output_file)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    fill_summary = PatternFill("solid", fgColor="D9EAF7")
    fill_conflict = PatternFill("solid", fgColor="FCE5CD")
    fills = {
        "EXACT": PatternFill("solid", fgColor=normalize_fill_color(result.config.get("color_exact"), DEFAULT_MATCH_COLORS["color_exact"])),
        "MATCH": PatternFill("solid", fgColor=normalize_fill_color(result.config.get("color_match"), DEFAULT_MATCH_COLORS["color_match"])),
        "REVIEW": PatternFill("solid", fgColor=normalize_fill_color(result.config.get("color_review"), DEFAULT_MATCH_COLORS["color_review"])),
        "NO_MATCH": PatternFill("solid", fgColor=normalize_fill_color(result.config.get("color_no_match"), DEFAULT_MATCH_COLORS["color_no_match"])),
        "EXCESS_LEFT": PatternFill("solid", fgColor=normalize_fill_color(result.config.get("color_excess_left"), DEFAULT_MATCH_COLORS["color_excess_left"])),
    }

    source_bucket_by_row = result.results_df.set_index("source_row_id")["final_color_bucket"].fillna("").to_dict()
    target_row_map = result.target_df.set_index("excel_row_t2")["target_row_id"].to_dict() if not result.target_df.empty else {}
    money_cols_t1 = parse_csv_columns(result.config.get("tab4_money_cols_t1", ""))
    money_cols_t2 = parse_csv_columns(result.config.get("tab4_money_cols_t2", ""))
    target_bucket_by_row_id: dict[int, str] = {
        int(row_id): "NO_MATCH" for row_id in result.target_df["target_row_id"].tolist()
    } if not result.target_df.empty else {}
    priority = {"EXACT": 5, "MATCH": 4, "REVIEW": 3, "EXCESS_LEFT": 2, "NO_MATCH": 1, "": 0}
    for row in result.results_df.itertuples(index=False):
        line_match = getattr(row, "final_line_match_t2", pd.NA)
        if pd.isna(line_match):
            continue
        target_row_id = target_row_map.get(int(line_match))
        if target_row_id is None:
            continue
        bucket = str(getattr(row, "final_color_bucket", "") or "")
        current = target_bucket_by_row_id.get(int(target_row_id), "")
        if priority.get(bucket, 0) >= priority.get(current, 0):
            target_bucket_by_row_id[int(target_row_id)] = bucket

    for ws in wb.worksheets:
        if ws.title == "resultados_match":
            summary_header_row = 1
            data_header_row = results_startrow + 1
            for cell in ws[summary_header_row]:
                cell.fill = header_fill
                cell.font = header_font
            for row in ws.iter_rows(min_row=2, max_row=results_startrow):
                for cell in row:
                    if cell.value not in (None, ""):
                        cell.fill = fill_summary
            for cell in ws[data_header_row]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = f"A{data_header_row + 1}"
            for excel_row, bucket in enumerate(result.results_df["final_color_bucket"].fillna("").tolist(), start=data_header_row + 1):
                fill = fills.get(str(bucket), None)
                if fill is not None:
                    for cell in ws[excel_row]:
                        cell.fill = fill
            conflict_col = _find_header_index(ws, "final_conflict_flags")
            if conflict_col is not None:
                for row in ws.iter_rows(min_row=data_header_row + 1, max_row=ws.max_row):
                    value = str(row[conflict_col - 1].value or "").strip()
                    if value:
                        row[conflict_col - 1].fill = fill_conflict
        elif ws.title == "excel_1_original":
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = "A2"
            for source_row_id, bucket in source_bucket_by_row.items():
                fill = fills.get(str(bucket), None)
                if fill is None:
                    continue
                excel_row = int(source_row_id) + 1
                for cell in ws[excel_row]:
                    cell.fill = fill
        elif ws.title == "excel_2_original":
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = "A2"
            for target_row_id, bucket in target_bucket_by_row_id.items():
                fill = fills.get(str(bucket), None)
                if fill is None:
                    continue
                excel_row = int(target_row_id) + 1
                for cell in ws[excel_row]:
                    cell.fill = fill
        elif ws.title == "conciliacao_quantidades":
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws.freeze_panes = "A2"
            reconciliation_df = result.grouped_reconciliation_df
            for row_number, row in enumerate(reconciliation_df.to_dict("records"), start=2):
                if row.get("_bucket_left"):
                    ws.cell(row=row_number, column=1).fill = fills.get(str(row["_bucket_left"]), PatternFill())
                if row.get("_bucket_right"):
                    ws.cell(row=row_number, column=2).fill = fills.get(str(row["_bucket_right"]), PatternFill())
            money_headers = [f"E1:{column}" for column in money_cols_t1] + [f"E2:{column}" for column in money_cols_t2]
            for header_name in money_headers:
                header_index = _find_header_index(ws, header_name)
                if header_index is None:
                    continue
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row[header_index - 1]
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
        ws.auto_filter.ref = ws.dimensions
        _autosize_columns(ws)

    wb.save(output_file)


def open_file_with_default_app(path: Path) -> None:
    path = path.resolve()
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


@dataclass(frozen=True)
class WorkbookMetadata:
    file_path: Path
    sheet_names: list[str]


def list_workbook_sheets(file_path: str | Path, errors: list[str] | None = None) -> list[str]:
    path = Path(file_path)
    if not path.exists():
        return []
    try:
        return list(pd.ExcelFile(path).sheet_names)
    except Exception as exc:
        if errors is not None:
            errors.append(
                "Falha ao ler abas da planilha "
                f"'{path}': {exc}. Verifique se o arquivo não está corrompido, bloqueado ou protegido."
            )
        return []


def read_workbook_metadata(file_path: str | Path, errors: list[str] | None = None) -> WorkbookMetadata:
    path = Path(file_path)
    sheet_names = list_workbook_sheets(path, errors=errors)
    return WorkbookMetadata(file_path=path, sheet_names=sheet_names)


def normalize_fill_color(value: Any, default: str) -> str:
    resolved = resolve_fill_color(value)
    if resolved:
        return resolved
    fallback = resolve_fill_color(default)
    return fallback or "FF000000"


def canonicalize_color_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return re.sub(r"\s+", " ", text)


def build_color_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for label, fill_code in COLOR_NAME_TO_FILL.items():
        aliases[canonicalize_color_name(label)] = fill_code

    semantic_aliases = {
        "exact": "Verde (Exato)",
        "match": "Azul (Match aceito)",
        "review": "Laranja (Revisão)",
        "no match": "Vermelho (Sem match)",
        "excess left": "Azul claro (Excedente Excel 1)",
        "excedente excel 1": "Azul claro (Excedente Excel 1)",
        "sem match": "Vermelho (Sem match)",
        "revisao": "Laranja (Revisão)",
        "aceito": "Azul (Match aceito)",
        "exato": "Verde (Exato)",
        "verde": "Verde (Exato)",
        "azul": "Azul (Match aceito)",
        "laranja": "Laranja (Revisão)",
        "vermelho": "Vermelho (Sem match)",
        "azul claro": "Azul claro (Excedente Excel 1)",
    }
    for alias, label in semantic_aliases.items():
        aliases[canonicalize_color_name(alias)] = COLOR_NAME_TO_FILL[label]

    for key, fill_code in DEFAULT_MATCH_COLORS.items():
        aliases[canonicalize_color_name(key)] = fill_code
    return aliases


COLOR_ALIAS_TO_FILL = build_color_alias_map()


def resolve_fill_color(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    hex_candidate = text.replace("#", "").strip().upper()
    if re.fullmatch(r"[0-9A-F]{6}", hex_candidate):
        return f"FF{hex_candidate}"
    if re.fullmatch(r"[0-9A-F]{8}", hex_candidate):
        return hex_candidate
    return COLOR_ALIAS_TO_FILL.get(canonicalize_color_name(text))


def format_color_for_ui(value: Any, default_fill: str) -> str:
    resolved = resolve_fill_color(value)
    if resolved:
        return COLOR_FILL_TO_LABEL.get(resolved, f"#{resolved[-6:]}")
    text = str(value or "").strip()
    if text:
        return text
    default_resolved = normalize_fill_color(default_fill, default_fill)
    return COLOR_FILL_TO_LABEL.get(default_resolved, f"#{default_resolved[-6:]}")


def parse_csv_columns(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def serialize_csv_columns(columns: list[str]) -> str:
    return ", ".join([str(column).strip() for column in columns if str(column).strip()])


def parse_brl_currency_value(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()

    text = text.replace("R$", "").replace("\xa0", " ")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text:
        return None

    if text.count("-") > 1:
        return None
    if "-" in text:
        if text.startswith("-"):
            negative = True
            text = text[1:]
        else:
            return None
    if not text:
        return None

    if "," in text:
        text = text.replace(".", "")
        text = text.replace(",", ".")

    try:
        numeric = float(text)
    except ValueError:
        return None
    return -numeric if negative else numeric


def list_workbook_columns(
    file_path: str | Path,
    sheet_name: str,
    header_row: int,
    errors: list[str] | None = None,
) -> list[str]:
    path = Path(file_path)
    if not path.exists():
        return []
    try:
        preview_df = pd.read_excel(
            path,
            sheet_name=sheet_name,
            header=max(int(header_row) - 1, 0),
            dtype=str,
            nrows=0,
        )
        return [str(col) for col in preview_df.columns]
    except Exception as exc:
        if errors is not None:
            errors.append(
                "Falha ao ler colunas da planilha "
                f"'{path}' (aba='{sheet_name}', cabeçalho={header_row}): {exc}. "
                "Verifique se a aba/linha de cabeçalho está correta e se o arquivo está acessível."
            )
        return []


def collect_workbook_preview(config: dict[str, Any]) -> str:
    lines: list[str] = []
    for label, file_key, sheet_key, header_key, col_key, match2_key in (
        ("Excel 1", "input_file_t1", "sheet_t1", "header_row_t1", "name_col_t1", "match2_col_t1"),
        ("Excel 2", "input_file_t2", "sheet_t2", "header_row_t2", "name_col_t2", "match2_col_t2"),
    ):
        input_file = Path(config[file_key])
        if not input_file.exists():
            raise FileNotFoundError(f"Arquivo de entrada não encontrado: {input_file}")

        xls = pd.ExcelFile(input_file)
        lines.extend(
            [
                f"{label}: {input_file}",
                f"  - Quantidade de abas: {len(xls.sheet_names)}",
                f"  - Abas: {', '.join(xls.sheet_names)}",
            ]
        )
        sheet_name = config[sheet_key]
        header_row = int(config[header_key])
        name_col = config[col_key]
        if sheet_name not in xls.sheet_names:
            lines.append(f"  - Aba ausente: {sheet_name}")
            lines.append("")
            continue

        preview_df = pd.read_excel(
            input_file,
            sheet_name=sheet_name,
            header=header_row - 1,
            dtype=str,
            nrows=5,
        )
        headers = [str(col) for col in preview_df.columns]
        lines.append(f"  - Aba ativa: {sheet_name}")
        lines.append(f"  - Linha de cabeçalho: {header_row}")
        lines.append(f"  - Colunas: {', '.join(headers[:15])}")
        selected_index = excel_col_to_index(name_col)
        if selected_index >= len(headers):
            lines.append(f"  - A coluna de nome {name_col} está fora do intervalo")
        else:
            lines.append(f"  - Coluna de nome {name_col} -> {headers[selected_index]}")
            sample_values = preview_df.iloc[:, selected_index].fillna("").astype(str).head(5).tolist()
            lines.append(f"  - Valores de exemplo: {sample_values}")
        match2_col = str(config.get(match2_key, "") or "").strip()
        if match2_col:
            match2_index = excel_col_to_index(match2_col)
            if match2_index >= len(headers):
                lines.append(f"  - Coluna match 2 {match2_col} fora do intervalo")
            else:
                lines.append(f"  - Coluna match 2 {match2_col} -> {headers[match2_index]}")
                sample_match2 = preview_df.iloc[:, match2_index].fillna("").astype(str).head(5).tolist()
                lines.append(f"  - Match 2 (amostra): {sample_match2}")
        lines.append("")
    lines.append(
        f"Match 2: prefixo={config.get('match2_prefix_chars', 8)} chars, peso auxiliar={config.get('match2_weight', 0.2)}"
    )
    lines.append(f"Match 1: limite efetivo de caracteres = {config.get('max_external_chars', 30)}")
    lines.append(
        f"Extras aba 4 E1: {config.get('tab4_extra_cols_t1', '') or '(nenhum)'} | E2: {config.get('tab4_extra_cols_t2', '') or '(nenhum)'}"
    )
    lines.append(
        f"Monetárias aba 4 E1: {config.get('tab4_money_cols_t1', '') or '(nenhum)'} | E2: {config.get('tab4_money_cols_t2', '') or '(nenhum)'}"
    )
    return "\n".join(lines).strip()


def normalize_config_values(config: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(config)
    normalized["input_file_t1"] = str(Path(normalized["input_file_t1"]).expanduser())
    normalized["input_file_t2"] = str(Path(normalized["input_file_t2"]).expanduser())
    normalized["output_file"] = str(Path(normalized["output_file"]).expanduser())

    required_text = [
        "input_file_t1",
        "input_file_t2",
        "output_file",
        "sheet_t1",
        "sheet_t2",
        "name_col_t1",
        "name_col_t2",
        "output_mode",
        "quantity_resolution_mode",
    ]
    for key in required_text:
        if not str(normalized.get(key, "")).strip():
            raise ValueError(f"O campo '{key}' não pode ficar vazio.")

    normalized["header_row_t1"] = int(normalized["header_row_t1"])
    normalized["header_row_t2"] = int(normalized["header_row_t2"])
    normalized["max_external_chars"] = int(normalized["max_external_chars"])
    normalized["top_candidates_to_keep"] = int(normalized["top_candidates_to_keep"])
    normalized["max_matches_per_t2_name"] = int(normalized["max_matches_per_t2_name"])
    normalized["accept_score"] = float(normalized["accept_score"])
    normalized["review_score"] = float(normalized["review_score"])
    normalized["min_gap_for_accept"] = float(normalized["min_gap_for_accept"])
    normalized["length_gap_penalty_per_char"] = float(normalized.get("length_gap_penalty_per_char", 0.5))
    normalized["max_length_gap_penalty"] = float(normalized.get("max_length_gap_penalty", 10.0))
    normalized["missing_surname_penalty"] = float(normalized.get("missing_surname_penalty", 3.0))
    normalized["match2_prefix_chars"] = int(normalized.get("match2_prefix_chars", 8) or 8)
    normalized["match2_weight"] = float(normalized.get("match2_weight", 0.2) or 0.2)
    normalized["allow_reuse_t2_matches"] = parse_bool(normalized["allow_reuse_t2_matches"], "allow_reuse_t2_matches")
    normalized["auto_open_output"] = parse_bool(normalized["auto_open_output"], "auto_open_output")
    normalized["match2_col_t1"] = str(normalized.get("match2_col_t1", "") or "").strip().upper()
    normalized["match2_col_t2"] = str(normalized.get("match2_col_t2", "") or "").strip().upper()
    normalized["tab4_extra_cols_t1"] = serialize_csv_columns(parse_csv_columns(normalized.get("tab4_extra_cols_t1", "")))
    normalized["tab4_extra_cols_t2"] = serialize_csv_columns(parse_csv_columns(normalized.get("tab4_extra_cols_t2", "")))
    normalized["tab4_money_cols_t1"] = serialize_csv_columns(parse_csv_columns(normalized.get("tab4_money_cols_t1", "")))
    normalized["tab4_money_cols_t2"] = serialize_csv_columns(parse_csv_columns(normalized.get("tab4_money_cols_t2", "")))
    for key, default_value in DEFAULT_SCORE_WEIGHTS.items():
        normalized[key] = float(normalized.get(key, default_value))
    for key, default_value in DEFAULT_MATCH_COLORS.items():
        normalized[key] = normalize_fill_color(normalized.get(key), default_value)
    return normalized


def validate_config_thresholds_and_columns(normalized: dict[str, Any]) -> None:
    if normalized["header_row_t1"] <= 0 or normalized["header_row_t2"] <= 0:
        raise ValueError("As linhas de cabeçalho devem ser maiores que zero.")
    if normalized["max_external_chars"] <= 0:
        raise ValueError("O tamanho do prefixo deve ser maior que zero.")
    if normalized["top_candidates_to_keep"] <= 0:
        raise ValueError("A quantidade de candidatos para manter deve ser maior que zero.")
    if normalized["max_matches_per_t2_name"] <= 0:
        raise ValueError("O limite de reaproveitamento deve ser maior que zero.")
    if normalized["accept_score"] < normalized["review_score"]:
        raise ValueError("A pontuação de aceite deve ser maior ou igual à pontuação de revisão.")
    if normalized["accept_score"] > 100 or normalized["review_score"] > 100:
        raise ValueError("As pontuações devem ficar entre 0 e 100.")
    if normalized["length_gap_penalty_per_char"] < 0:
        raise ValueError("A penalidade por diferença de tamanho por caractere deve ser zero ou maior.")
    if normalized["max_length_gap_penalty"] < 0:
        raise ValueError("A penalidade máxima por diferença de tamanho deve ser zero ou maior.")
    if normalized["missing_surname_penalty"] < 0:
        raise ValueError("A penalidade por sobrenome ausente deve ser zero ou maior.")
    if normalized["match2_prefix_chars"] <= 0:
        raise ValueError("A quantidade de caracteres do match 2 deve ser maior que zero.")
    if normalized["match2_weight"] < 0:
        raise ValueError("O peso do match 2 deve ser zero ou maior.")
    if all(normalized[key] <= 0 for key in DEFAULT_SCORE_WEIGHTS):
        raise ValueError("Pelo menos um peso de pontuação deve ser maior que zero.")
    if bool(normalized["match2_col_t1"]) != bool(normalized["match2_col_t2"]):
        raise ValueError("Para usar match 2, configure a coluna em ambos os excels.")

    excel_col_to_index(normalized["name_col_t1"])
    excel_col_to_index(normalized["name_col_t2"])
    if normalized["match2_col_t1"]:
        excel_col_to_index(normalized["match2_col_t1"])
    if normalized["match2_col_t2"]:
        excel_col_to_index(normalized["match2_col_t2"])


def validate_config_workbook_metadata(normalized: dict[str, Any]) -> None:
    file_t1 = Path(normalized["input_file_t1"])
    file_t2 = Path(normalized["input_file_t2"])
    if not file_t1.exists():
        raise FileNotFoundError(f"Arquivo de entrada não encontrado: {file_t1}")
    if not file_t2.exists():
        raise FileNotFoundError(f"Arquivo de entrada não encontrado: {file_t2}")
    metadata_errors: list[str] = []
    metadata_t1 = read_workbook_metadata(file_t1, errors=metadata_errors)
    metadata_t2 = read_workbook_metadata(file_t2, errors=metadata_errors)
    if metadata_errors:
        raise ValueError("\n".join(metadata_errors))
    if normalized["sheet_t1"] not in metadata_t1.sheet_names:
        raise ValueError(f"Aba ausente no Excel 1: {normalized['sheet_t1']}")
    if normalized["sheet_t2"] not in metadata_t2.sheet_names:
        raise ValueError(f"Aba ausente no Excel 2: {normalized['sheet_t2']}")
    headers_t1 = list_workbook_columns(
        metadata_t1.file_path,
        normalized["sheet_t1"],
        normalized["header_row_t1"],
        errors=metadata_errors,
    )
    headers_t2 = list_workbook_columns(
        metadata_t2.file_path,
        normalized["sheet_t2"],
        normalized["header_row_t2"],
        errors=metadata_errors,
    )
    if metadata_errors:
        raise ValueError("\n".join(metadata_errors))
    if normalized["match2_col_t1"] and excel_col_to_index(normalized["match2_col_t1"]) >= len(headers_t1):
        raise ValueError(f"Coluna match 2 inválida no Excel 1: {normalized['match2_col_t1']}")
    if normalized["match2_col_t2"] and excel_col_to_index(normalized["match2_col_t2"]) >= len(headers_t2):
        raise ValueError(f"Coluna match 2 inválida no Excel 2: {normalized['match2_col_t2']}")
    extras_t1 = parse_csv_columns(normalized["tab4_extra_cols_t1"])
    extras_t2 = parse_csv_columns(normalized["tab4_extra_cols_t2"])
    money_t1 = parse_csv_columns(normalized["tab4_money_cols_t1"])
    money_t2 = parse_csv_columns(normalized["tab4_money_cols_t2"])
    missing_t1 = [col for col in extras_t1 if col not in headers_t1]
    missing_t2 = [col for col in extras_t2 if col not in headers_t2]
    missing_money_t1 = [col for col in money_t1 if col not in headers_t1]
    missing_money_t2 = [col for col in money_t2 if col not in headers_t2]
    missing_money_in_extras_t1 = [col for col in money_t1 if col not in extras_t1]
    missing_money_in_extras_t2 = [col for col in money_t2 if col not in extras_t2]
    if missing_t1:
        raise ValueError(f"Coluna(s) extra(s) da aba 4 ausente(s) no Excel 1: {', '.join(missing_t1)}")
    if missing_t2:
        raise ValueError(f"Coluna(s) extra(s) da aba 4 ausente(s) no Excel 2: {', '.join(missing_t2)}")
    if missing_money_t1:
        raise ValueError(f"Coluna(s) monetária(s) da aba 4 ausente(s) no Excel 1: {', '.join(missing_money_t1)}")
    if missing_money_t2:
        raise ValueError(f"Coluna(s) monetária(s) da aba 4 ausente(s) no Excel 2: {', '.join(missing_money_t2)}")
    if missing_money_in_extras_t1:
        raise ValueError(
            "Coluna(s) monetária(s) da aba 4 no Excel 1 devem também estar em Extras aba 4: "
            + ", ".join(missing_money_in_extras_t1)
        )
    if missing_money_in_extras_t2:
        raise ValueError(
            "Coluna(s) monetária(s) da aba 4 no Excel 2 devem também estar em Extras aba 4: "
            + ", ".join(missing_money_in_extras_t2)
        )


def validate_config(config: dict[str, Any], validate_workbook: bool = True) -> dict[str, Any]:
    normalized = normalize_config_values(config)
    validate_config_thresholds_and_columns(normalized)
    if validate_workbook:
        validate_config_workbook_metadata(normalized)

    return normalized


def prepare_input_frames(config: dict[str, Any], progress_callback: ProgressCallback | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    input_file_t1 = Path(config["input_file_t1"])
    input_file_t2 = Path(config["input_file_t2"])
    emit_progress(progress_callback, "Lendo abas das planilhas...", 5)
    df1 = pd.read_excel(
        input_file_t1,
        sheet_name=config["sheet_t1"],
        header=config["header_row_t1"] - 1,
        dtype=str,
    )
    df2 = pd.read_excel(
        input_file_t2,
        sheet_name=config["sheet_t2"],
        header=config["header_row_t2"] - 1,
        dtype=str,
    )

    idx_t1 = excel_col_to_index(config["name_col_t1"])
    idx_t2 = excel_col_to_index(config["name_col_t2"])
    idx_match2_t1 = excel_col_to_index(config["match2_col_t1"]) if config.get("match2_col_t1") else None
    idx_match2_t2 = excel_col_to_index(config["match2_col_t2"]) if config.get("match2_col_t2") else None
    if idx_t1 >= len(df1.columns):
        raise IndexError(f"A coluna {config['name_col_t1']} não existe em {config['sheet_t1']}.")
    if idx_t2 >= len(df2.columns):
        raise IndexError(f"A coluna {config['name_col_t2']} não existe em {config['sheet_t2']}.")
    if idx_match2_t1 is not None and idx_match2_t1 >= len(df1.columns):
        raise IndexError(f"A coluna match 2 {config['match2_col_t1']} não existe em {config['sheet_t1']}.")
    if idx_match2_t2 is not None and idx_match2_t2 >= len(df2.columns):
        raise IndexError(f"A coluna match 2 {config['match2_col_t2']} não existe em {config['sheet_t2']}.")

    emit_progress(progress_callback, "Normalizando nomes e metadados...", 12)
    df1 = df1.copy()
    df2 = df2.copy()

    df1["source_row_id"] = df1.index + 1
    df2["target_row_id"] = df2.index + 1
    df1["excel_row_t1"] = config["header_row_t1"] + df1.index + 1
    df2["excel_row_t2"] = config["header_row_t2"] + df2.index + 1

    df1["nome_t1_original"] = df1.iloc[:, idx_t1].fillna("").astype(str)
    df2["nome_t2_original"] = df2.iloc[:, idx_t2].fillna("").astype(str)
    if idx_match2_t1 is not None:
        df1["match2_t1_original"] = df1.iloc[:, idx_match2_t1].fillna("").astype(str)
    else:
        df1["match2_t1_original"] = ""
    if idx_match2_t2 is not None:
        df2["match2_t2_original"] = df2.iloc[:, idx_match2_t2].fillna("").astype(str)
    else:
        df2["match2_t2_original"] = ""
    df1["nome_t1_norm"] = df1["nome_t1_original"].apply(normalize_name)
    df2["nome_t2_norm"] = df2["nome_t2_original"].apply(normalize_name)
    limit_chars = int(config["max_external_chars"])
    df1["nome_t1_match_norm"] = df1["nome_t1_norm"].str[:limit_chars]
    df2["nome_t2_match_norm"] = df2["nome_t2_norm"].str[:limit_chars]
    df1["match2_t1_norm"] = df1["match2_t1_original"].fillna("").astype(str).str.strip().str.upper()
    df2["match2_t2_norm"] = df2["match2_t2_original"].fillna("").astype(str).str.strip().str.upper()
    prefix_chars = int(config.get("match2_prefix_chars", 8))
    df1["match2_t1_prefix"] = df1["match2_t1_norm"].str[:prefix_chars]
    df2["match2_t2_prefix"] = df2["match2_t2_norm"].str[:prefix_chars]
    df1["first_token"] = df1["nome_t1_match_norm"].apply(first_token)
    df1["last_token"] = df1["nome_t1_match_norm"].apply(last_token)
    df2["first_token"] = df2["nome_t2_match_norm"].apply(first_token)
    df2["last_token"] = df2["nome_t2_match_norm"].apply(last_token)
    df1["key_prefix"] = df1["nome_t1_match_norm"]
    df2["key_prefix"] = df2["nome_t2_match_norm"]

    return df1, df2


def build_target_catalog(df2: pd.DataFrame, config: dict[str, Any]) -> tuple[pd.DataFrame, dict[str, list[dict[str, Any]]]]:
    df2_valid = df2[df2["nome_t2_match_norm"] != ""].copy()
    if df2_valid.empty:
        return pd.DataFrame(), {"by_prefix": defaultdict(list), "by_first": defaultdict(list), "by_last": defaultdict(list), "all": []}

    grouped = (
        df2_valid.sort_values("excel_row_t2")
        .groupby("nome_t2_match_norm", as_index=False)
        .agg(
            nome_t2_norm_full=("nome_t2_norm", "first"),
            nome_t2_original=("nome_t2_original", "first"),
            excel_row_t2=("excel_row_t2", "min"),
            first_token=("first_token", "first"),
            last_token=("last_token", "first"),
            key_prefix=("key_prefix", "first"),
            match2_t2_original=("match2_t2_original", "first"),
            match2_t2_norm=("match2_t2_norm", "first"),
            match2_t2_prefix=("match2_t2_prefix", "first"),
            quota_original=("nome_t2_match_norm", "size"),
        )
    )
    grouped = grouped.rename(columns={"nome_t2_match_norm": "nome_t2_norm"})
    if config["allow_reuse_t2_matches"]:
        grouped["quota_limit"] = int(config["max_matches_per_t2_name"])
    else:
        grouped["quota_limit"] = grouped["quota_original"]

    grouped["quota_limit"] = grouped["quota_limit"].astype(int)
    grouped["quota_original"] = grouped["quota_original"].astype(int)

    indexes: dict[str, Any] = {
        "by_prefix": defaultdict(list),
        "by_first": defaultdict(list),
        "by_last": defaultdict(list),
        "all": [],
    }

    records = grouped.to_dict("records")
    indexes["all"] = records
    for record in records:
        if record["key_prefix"]:
            indexes["by_prefix"][record["key_prefix"]].append(record)
        if record["first_token"]:
            indexes["by_first"][record["first_token"]].append(record)
        if record["last_token"]:
            indexes["by_last"][record["last_token"]].append(record)

    return grouped, indexes


def choose_candidate_pool(
    name_norm: str,
    key_prefix: str,
    ft: str,
    lt: str,
    target_indexes: dict[str, Any],
) -> list[dict[str, Any]]:
    seen: dict[str, dict[str, Any]] = {}

    def absorb(records: list[dict[str, Any]]) -> None:
        for record in records:
            seen.setdefault(record["nome_t2_norm"], record)

    if key_prefix:
        absorb(target_indexes["by_prefix"].get(key_prefix, []))
    if ft:
        absorb(target_indexes["by_first"].get(ft, []))
    if lt:
        absorb(target_indexes["by_last"].get(lt, []))

    if not seen:
        absorb(target_indexes["all"])
    return list(seen.values())


def candidate_utility(candidate: dict[str, Any]) -> int:
    utility = int(round(candidate.get("score_composite", candidate["score"]) * 100))
    if candidate["exact_norm"]:
        utility += 5000
    if candidate["exact_prefix"]:
        utility += 3000
    if candidate["same_last"]:
        utility += 300
    if candidate["ext_subset_in_full"] or candidate["full_subset_in_ext"]:
        utility += 250
    if candidate["score_ordered_chars"] >= 90:
        utility += 250
    if candidate["score_aligned_chars"] >= 85:
        utility += 180
    utility += max(0, 100 - int(candidate["rank"])) * 5
    return utility


def add_graph_edge(graph: list[list[FlowEdge]], source: int, target: int, capacity: int, cost: int) -> FlowEdge:
    forward = FlowEdge(target, len(graph[target]), capacity, cost)
    backward = FlowEdge(source, len(graph[source]), 0, -cost)
    graph[source].append(forward)
    graph[target].append(backward)
    return forward


def shortest_path_spfa(graph: list[list[FlowEdge]], source: int, sink: int) -> tuple[list[float], list[int], list[int]]:
    distance = [float("inf")] * len(graph)
    in_queue = [False] * len(graph)
    prev_node = [-1] * len(graph)
    prev_edge = [-1] * len(graph)

    distance[source] = 0
    queue: deque[int] = deque([source])
    in_queue[source] = True

    while queue:
        node = queue.popleft()
        in_queue[node] = False
        for edge_index, edge in enumerate(graph[node]):
            if edge.capacity <= 0:
                continue
            new_distance = distance[node] + edge.cost
            if new_distance < distance[edge.to_node]:
                distance[edge.to_node] = new_distance
                prev_node[edge.to_node] = node
                prev_edge[edge.to_node] = edge_index
                if not in_queue[edge.to_node]:
                    queue.append(edge.to_node)
                    in_queue[edge.to_node] = True
    return distance, prev_node, prev_edge


def solve_global_assignment(candidates_df: pd.DataFrame, quota_limits: dict[str, int]) -> pd.DataFrame:
    eligible = candidates_df[candidates_df["eligible_for_global"]].copy()
    if eligible.empty:
        return pd.DataFrame(columns=["source_row_id", "nome_t2_norm", "assignment_utility"])

    row_ids = sorted(eligible["source_row_id"].unique().tolist())
    targets = sorted(eligible["nome_t2_norm"].unique().tolist())
    row_node = {row_id: index + 1 for index, row_id in enumerate(row_ids)}
    target_node = {name: len(row_ids) + index + 1 for index, name in enumerate(targets)}
    sink = len(row_ids) + len(targets) + 1
    graph: list[list[FlowEdge]] = [[] for _ in range(sink + 1)]

    for row_id in row_ids:
        add_graph_edge(graph, 0, row_node[row_id], 1, 0)
    for target_name in targets:
        add_graph_edge(graph, target_node[target_name], sink, int(quota_limits.get(target_name, 0)), 0)

    assignment_edges: dict[tuple[int, str], FlowEdge] = {}
    for candidate in eligible.to_dict("records"):
        edge = add_graph_edge(
            graph,
            row_node[int(candidate["source_row_id"])],
            target_node[str(candidate["nome_t2_norm"])],
            1,
            -int(candidate["utility"]),
        )
        assignment_edges[(int(candidate["source_row_id"]), str(candidate["nome_t2_norm"]))] = edge

    while True:
        distance, prev_node, prev_edge = shortest_path_spfa(graph, 0, sink)
        if distance[sink] == float("inf") or distance[sink] >= 0:
            break

        flow = 1
        node = sink
        while node != 0:
            edge = graph[prev_node[node]][prev_edge[node]]
            flow = min(flow, edge.capacity)
            node = prev_node[node]

        node = sink
        while node != 0:
            edge = graph[prev_node[node]][prev_edge[node]]
            edge.capacity -= flow
            reverse = graph[edge.to_node][edge.rev_index]
            reverse.capacity += flow
            node = prev_node[node]

    assigned_rows: list[dict[str, Any]] = []
    for (row_id, target_name), edge in assignment_edges.items():
        if edge.capacity == 0:
            assigned_rows.append(
                {
                    "source_row_id": row_id,
                    "nome_t2_norm": target_name,
                    "assignment_utility": -edge.cost,
                }
            )
    return pd.DataFrame(assigned_rows)


def initialize_result_columns(results_df: pd.DataFrame, top_candidates_to_keep: int) -> None:
    for rank in range(1, top_candidates_to_keep + 1):
        results_df[f"cand_{rank}_nome"] = ""
        results_df[f"cand_{rank}_score"] = pd.NA

    defaults = {
        "analysis_status": "",
        "analysis_method": "",
        "analysis_match_t2_original": "",
        "analysis_match_t2_norm": "",
        "analysis_line_match_t2": pd.NA,
        "analysis_score": pd.NA,
        "analysis_score_composite": pd.NA,
        "analysis_score_gap": pd.NA,
        "analysis_conflict_flags": "",
        "analysis_review_reason": "",
        "analysis_match2_t1": "",
        "analysis_match2_t2": "",
        "analysis_match2_equal_prefix": False,
        "analysis_match2_score": pd.NA,
        "manual_status": "",
        "manual_match_t2_original": "",
        "manual_match_t2_norm": "",
        "manual_line_match_t2": pd.NA,
        "manual_score": pd.NA,
        "manual_note": "",
        "manual_sequence": pd.NA,
        "final_status": "",
        "final_method": "",
        "final_match_t2_original": "",
        "final_match_t2_norm": "",
        "final_group_t2_original": "",
        "final_group_t2_norm": "",
        "final_line_match_t2": pd.NA,
        "final_score": pd.NA,
        "final_score_composite": pd.NA,
        "final_score_gap": pd.NA,
        "final_score_ordered_chars": pd.NA,
        "final_score_aligned_chars": pd.NA,
        "final_name_length_gap": pd.NA,
        "final_same_name_length": pd.NA,
        "final_color_bucket": "",
        "final_conflict_flags": "",
        "final_review_required": False,
        "final_match2_t1": "",
        "final_match2_t2": "",
        "final_match2_equal_prefix": False,
        "final_match2_score": pd.NA,
        "final_quota_limit": pd.NA,
        "final_quota_order": pd.NA,
        "final_within_quota": pd.NA,
    }
    for column_name, default_value in defaults.items():
        results_df[column_name] = default_value


def analyze_matching(config: dict[str, Any], progress_callback: ProgressCallback | None = None) -> AnalysisResult:
    config = validate_config(config, validate_workbook=True)
    emit_progress(progress_callback, "Preparando sessão de análise...", 0)
    df1, df2 = prepare_input_frames(config, progress_callback)
    catalog_df, target_indexes = build_target_catalog(df2, config)

    results_df = df1.copy()
    initialize_result_columns(results_df, config["top_candidates_to_keep"])

    if catalog_df.empty:
        results_df["analysis_status"] = "SEM_MATCH"
        results_df["analysis_method"] = "SEM_TABELA2"
        results_df["analysis_review_reason"] = "A Tabela 2 não contém nomes normalizados."
        recompute_final_state(results_df, pd.DataFrame(columns=catalog_df.columns), config=config)
        summary_df = build_summary(results_df)
        return AnalysisResult(
            config=config,
            results_df=results_df,
            candidates_df=pd.DataFrame(),
            catalog_df=pd.DataFrame(),
            quota_df=pd.DataFrame(),
            summary_df=summary_df,
            review_df=results_df[results_df["final_status"] == "REVISAR"].copy(),
            preview_df=results_df.head(30).copy(),
            source_df=df1.copy(),
            target_df=df2.copy(),
            grouped_reconciliation_df=pd.DataFrame(),
        )

    candidate_rows: list[dict[str, Any]] = []
    emit_progress(progress_callback, "Calculando pontuação dos grupos de candidatos...", 18)
    rows = list(results_df.index)
    internal_keep = max(config["top_candidates_to_keep"], 8)

    for position, row_index in enumerate(rows, start=1):
        if position % 50 == 0 or position == len(rows):
            percent = 18 + (position / max(len(rows), 1)) * 40
            emit_progress(progress_callback, f"Pontuando candidatos {position}/{len(rows)}...", percent)

        row = results_df.loc[row_index]
        source_row_id = int(row["source_row_id"])
        name_norm = row["nome_t1_match_norm"]
        if not name_norm:
            results_df.at[row_index, "analysis_status"] = "SEM_MATCH"
            results_df.at[row_index, "analysis_method"] = "SEM_DADO"
            results_df.at[row_index, "analysis_review_reason"] = "Nome de entrada em branco."
            continue

        pool = choose_candidate_pool(
            name_norm,
            str(row["key_prefix"]),
            str(row["first_token"]),
            str(row["last_token"]),
            target_indexes,
        )
        scored: list[dict[str, Any]] = []
        for record in pool:
            metrics = score_candidate(
                name_norm,
                str(record["nome_t2_norm"]),
                config["max_external_chars"],
                match2_left=str(row.get("match2_t1_norm", "") or ""),
                match2_right=str(record.get("match2_t2_norm", "") or ""),
                match2_prefix_chars=int(config.get("match2_prefix_chars", 8)),
                match2_weight=float(config.get("match2_weight", 0.2)),
                config=config,
            )
            exact_norm = name_norm == record["nome_t2_norm"]
            exact_prefix = bool(row["key_prefix"]) and str(row["key_prefix"]) == str(record["key_prefix"])
            scored.append(
                {
                    "source_row_id": source_row_id,
                    "excel_row_t1": row["excel_row_t1"],
                    "nome_t1_original": row["nome_t1_original"],
                    "nome_t1_norm": name_norm,
                    "nome_t1_norm_full": row["nome_t1_norm"],
                    "match2_t1_original": row.get("match2_t1_original", ""),
                    "match2_t1_norm": row.get("match2_t1_norm", ""),
                    "nome_t2_original": record["nome_t2_original"],
                    "nome_t2_norm": record["nome_t2_norm"],
                    "nome_t2_norm_full": record.get("nome_t2_norm_full", record["nome_t2_norm"]),
                    "match2_t2_original": record.get("match2_t2_original", ""),
                    "match2_t2_norm": record.get("match2_t2_norm", ""),
                    "excel_row_t2": record["excel_row_t2"],
                    "quota_original": record["quota_original"],
                    "quota_limit": record["quota_limit"],
                    "exact_norm": exact_norm,
                    "exact_prefix": exact_prefix,
                    **metrics,
                }
            )

        scored.sort(
            key=lambda item: (
                item["score_composite"],
                item["score"],
                item["exact_norm"],
                item["exact_prefix"],
                item["score_ordered_chars"],
                item["score_aligned_chars"],
                item["same_last"],
            ),
            reverse=True,
        )

        kept = scored[:internal_keep]
        for rank, candidate in enumerate(kept, start=1):
            next_score = kept[rank]["score"] if rank < len(kept) else 0.0
            next_composite = kept[rank]["score_composite"] if rank < len(kept) else 0.0
            gap_to_next = round(candidate["score_composite"] - next_composite, 2)
            candidate["rank"] = rank
            candidate["gap_to_next"] = gap_to_next
            candidate["review_eligible"] = bool(candidate["score_composite"] >= config["review_score"] and candidate["same_first"])
            candidate["eligible_for_global"] = bool(
                candidate["exact_norm"]
                or candidate["exact_prefix"]
                or (candidate["score_composite"] >= config["accept_score"] and candidate["structure_ok"])
            )
            candidate["confident_if_top"] = bool(
                rank == 1
                and (
                    candidate["exact_norm"]
                    or candidate["exact_prefix"]
                    or (
                        candidate["score_composite"] >= config["accept_score"]
                        and candidate["structure_ok"]
                        and gap_to_next >= config["min_gap_for_accept"]
                    )
                )
                and not candidate["needs_length_review"]
            )
            candidate["utility"] = candidate_utility(candidate)
            candidate_rows.append(candidate)

            if rank <= config["top_candidates_to_keep"]:
                results_df.at[row_index, f"cand_{rank}_nome"] = candidate["nome_t2_original"]
                results_df.at[row_index, f"cand_{rank}_score"] = candidate["score_composite"]

    candidates_df = pd.DataFrame(candidate_rows)
    emit_progress(progress_callback, "Executando alocação global com controle de cotas...", 62)

    quota_limits = (
        catalog_df.set_index("nome_t2_norm")["quota_limit"].astype(int).to_dict()
        if not catalog_df.empty
        else {}
    )
    assignments_df = solve_global_assignment(candidates_df, quota_limits)

    best_candidates = candidates_df[candidates_df["rank"] == 1].copy() if not candidates_df.empty else pd.DataFrame()
    assigned_candidates = (
        assignments_df.merge(
            candidates_df,
            on=["source_row_id", "nome_t2_norm"],
            how="left",
        )
        if not assignments_df.empty
        else pd.DataFrame()
    )

    best_map = {int(row["source_row_id"]): row for row in best_candidates.to_dict("records")}
    assigned_map = {int(row["source_row_id"]): row for row in assigned_candidates.to_dict("records")}
    eligible_counts = (
        candidates_df[candidates_df["eligible_for_global"]]
        .groupby("source_row_id")
        .size()
        .to_dict()
        if not candidates_df.empty
        else {}
    )

    emit_progress(progress_callback, "Classificando resultados da análise...", 78)
    for row_index in results_df.index:
        source_row_id = int(results_df.at[row_index, "source_row_id"])
        best = best_map.get(source_row_id)
        assigned = assigned_map.get(source_row_id)
        flags: list[str] = []

        if not results_df.at[row_index, "analysis_status"]:
            if not best:
                results_df.at[row_index, "analysis_status"] = "SEM_MATCH"
                results_df.at[row_index, "analysis_method"] = "SEM_CANDIDATO"
                results_df.at[row_index, "analysis_review_reason"] = "Nenhum candidato foi gerado."
            else:
                add_flag(flags, "LOW_GAP", best["gap_to_next"] < config["min_gap_for_accept"] and not (best["exact_norm"] or best["exact_prefix"]))
                add_flag(flags, "STRUCTURE_WARNING", best["score_composite"] >= config["accept_score"] and not best["structure_ok"])
                add_flag(flags, "QUOTA_CONFLICT", bool(eligible_counts.get(source_row_id)) and assigned is None)
                add_flag(flags, "GLOBAL_REALLOCATED", assigned is not None and int(assigned["rank"]) > 1)
                add_flag(flags, "LENGTH_POSITION_REVIEW", bool(best.get("needs_length_review", False)))

                chosen = assigned if assigned is not None else best
                results_df.at[row_index, "analysis_match_t2_original"] = chosen["nome_t2_original"]
                results_df.at[row_index, "analysis_match_t2_norm"] = chosen["nome_t2_norm"]
                results_df.at[row_index, "analysis_line_match_t2"] = chosen["excel_row_t2"]
                results_df.at[row_index, "analysis_score"] = chosen["score"]
                results_df.at[row_index, "analysis_score_composite"] = chosen["score_composite"]
                results_df.at[row_index, "analysis_score_gap"] = chosen["gap_to_next"]
                results_df.at[row_index, "analysis_match2_t1"] = str(results_df.at[row_index, "match2_t1_original"] or "")
                results_df.at[row_index, "analysis_match2_t2"] = chosen.get("match2_t2_original", "")
                results_df.at[row_index, "analysis_match2_equal_prefix"] = bool(chosen.get("match2_equal_prefix", False))
                results_df.at[row_index, "analysis_match2_score"] = chosen.get("match2_score", pd.NA)

                if assigned is not None and int(assigned["rank"]) == 1 and bool(assigned["confident_if_top"]):
                    results_df.at[row_index, "analysis_status"] = "ACEITO"
                    results_df.at[row_index, "analysis_method"] = (
                        "EXATO_GLOBAL" if assigned["exact_norm"] or assigned["exact_prefix"] else "FUZZY_GLOBAL"
                    )
                    results_df.at[row_index, "analysis_review_reason"] = ""
                elif assigned is not None:
                    results_df.at[row_index, "analysis_status"] = "REVISAR"
                    results_df.at[row_index, "analysis_method"] = "GLOBAL_REVIEW"
                    reasons = []
                    if int(assigned["rank"]) > 1:
                        reasons.append("A alocação global usou um candidato alternativo.")
                    if best["gap_to_next"] < config["min_gap_for_accept"]:
                        reasons.append("A diferença para o candidato seguinte ficou abaixo do limite de aceite automático.")
                    if bool(best.get("needs_length_review", False)):
                        reasons.append("A similaridade atingiu pontuação máxima, mas o tamanho do nome ou a posição dos caracteres não confere totalmente.")
                    results_df.at[row_index, "analysis_review_reason"] = " ".join(reasons) or "O candidato alternativo da alocação global deve ser revisado."
                elif bool(eligible_counts.get(source_row_id)):
                    results_df.at[row_index, "analysis_status"] = "REVISAR"
                    results_df.at[row_index, "analysis_method"] = "QUOTA_CONFLICT"
                    results_df.at[row_index, "analysis_review_reason"] = "Um candidato forte perdeu a cota na alocação global."
                elif bool(best.get("needs_length_review", False)):
                    results_df.at[row_index, "analysis_status"] = "REVISAR"
                    results_df.at[row_index, "analysis_method"] = "LENGTH_POSITION_REVIEW"
                    results_df.at[row_index, "analysis_review_reason"] = "A similaridade está muito alta, mas o tamanho ou o alinhamento de caracteres está incompleto."
                elif bool(best["review_eligible"]):
                    results_df.at[row_index, "analysis_status"] = "REVISAR"
                    results_df.at[row_index, "analysis_method"] = "FUZZY_REVIEW"
                    results_df.at[row_index, "analysis_review_reason"] = "O candidato precisa de confirmação manual."
                else:
                    results_df.at[row_index, "analysis_status"] = "SEM_MATCH"
                    results_df.at[row_index, "analysis_method"] = "SEM_MATCH"
                    results_df.at[row_index, "analysis_review_reason"] = "Nenhum candidato atingiu o limite mínimo para revisão."

        results_df.at[row_index, "analysis_conflict_flags"] = flags_to_text(flags)

    emit_progress(progress_callback, "Aplicando estado final padrão...", 88)
    recompute_final_state(results_df, catalog_df, config=config)
    summary_df = build_summary(results_df)
    review_df = results_df[results_df["final_status"] == "REVISAR"].copy()
    preview_columns = [
        "excel_row_t1",
        "nome_t1_original",
        "analysis_status",
        "final_status",
        "final_match_t2_original",
        "final_score",
        "final_conflict_flags",
    ]
    preview_df = results_df[[col for col in preview_columns if col in results_df.columns]].head(40).copy()
    emit_progress(progress_callback, "Análise concluída.", 100)
    return AnalysisResult(
        config=config,
        results_df=results_df,
        candidates_df=candidates_df,
        catalog_df=catalog_df,
        quota_df=build_quota_summary(results_df, catalog_df),
        summary_df=summary_df,
        review_df=review_df,
        preview_df=preview_df,
        source_df=df1.copy(),
        target_df=df2.copy(),
        grouped_reconciliation_df=pd.DataFrame(),
    )


def build_quota_summary(results_df: pd.DataFrame, catalog_df: pd.DataFrame) -> pd.DataFrame:
    if catalog_df.empty:
        return pd.DataFrame(columns=["nome_t2_norm", "nome_t2_original", "quota_original", "quota_limit", "accepted_count"])

    accepted_counts = (
        results_df.loc[
            (results_df["final_status"] == "ACEITO") & (results_df["final_match_t2_norm"] != ""),
            "final_match_t2_norm",
        ]
        .value_counts()
        .to_dict()
    )
    quota_df = catalog_df.copy()
    quota_df["accepted_count"] = quota_df["nome_t2_norm"].map(accepted_counts).fillna(0).astype(int)
    quota_df["remaining_quota"] = (quota_df["quota_limit"] - quota_df["accepted_count"]).astype(int)
    quota_df["is_full"] = quota_df["remaining_quota"] <= 0
    return quota_df.sort_values(["is_full", "nome_t2_original"], ascending=[False, True]).reset_index(drop=True)


def determine_color_bucket(
    final_status: str,
    final_score: Any,
    ordered_score: float,
    aligned_score: float,
    same_name_length: bool,
    final_method: str = "",
    source_norm: str = "",
    final_norm: str = "",
    match2_equal_prefix: bool = False,
    match2_active: bool = False,
    match2_values_present: bool = False,
    config: dict[str, Any] | None = None,
) -> str:
    if final_method == "EXCESS_LEFT":
        return "EXCESS_LEFT"
    if final_status == "SEM_MATCH":
        return "NO_MATCH"
    if final_status == "REVISAR":
        return "REVIEW"
    if final_status == "ACEITO":
        is_exact_name = bool(
            source_norm
            and final_norm
            and source_norm == final_norm
            and same_name_length
            and ordered_score >= 99.5
            and aligned_score >= 99.5
        )
        if is_exact_name and match2_active and match2_values_present and not match2_equal_prefix:
            return "MATCH"
        if is_exact_name:
            return "EXACT"
        return "MATCH"
    return "NO_MATCH"


def recompute_final_state(
    results_df: pd.DataFrame,
    catalog_df: pd.DataFrame,
    config: dict[str, Any] | None = None,
) -> None:
    quota_map = (
        catalog_df.set_index("nome_t2_norm")["quota_limit"].astype(int).to_dict()
        if not catalog_df.empty
        else {}
    )
    original_map = (
        catalog_df.set_index("nome_t2_norm")["nome_t2_original"].to_dict()
        if not catalog_df.empty
        else {}
    )
    line_map = (
        catalog_df.set_index("nome_t2_norm")["excel_row_t2"].to_dict()
        if not catalog_df.empty
        else {}
    )

    provisional: list[dict[str, Any]] = []
    for row_index in results_df.index:
        analysis_status = str(results_df.at[row_index, "analysis_status"] or "")
        analysis_method = str(results_df.at[row_index, "analysis_method"] or "")
        analysis_match_norm = str(results_df.at[row_index, "analysis_match_t2_norm"] or "")
        analysis_match_original = str(results_df.at[row_index, "analysis_match_t2_original"] or "")
        analysis_score = results_df.at[row_index, "analysis_score"]
        analysis_score_composite = results_df.at[row_index, "analysis_score_composite"]
        analysis_gap = results_df.at[row_index, "analysis_score_gap"]
        analysis_line = results_df.at[row_index, "analysis_line_match_t2"]
        analysis_flags = str(results_df.at[row_index, "analysis_conflict_flags"] or "")
        analysis_match2_t1 = str(results_df.at[row_index, "analysis_match2_t1"] or "")
        analysis_match2_t2 = str(results_df.at[row_index, "analysis_match2_t2"] or "")
        analysis_match2_equal_prefix = bool(results_df.at[row_index, "analysis_match2_equal_prefix"])
        analysis_match2_score = results_df.at[row_index, "analysis_match2_score"]

        manual_status = str(results_df.at[row_index, "manual_status"] or "").upper()
        manual_norm = str(results_df.at[row_index, "manual_match_t2_norm"] or "")
        manual_original = str(results_df.at[row_index, "manual_match_t2_original"] or "")
        manual_line = results_df.at[row_index, "manual_line_match_t2"]
        manual_score = results_df.at[row_index, "manual_score"]
        manual_note = str(results_df.at[row_index, "manual_note"] or "")

        final_status = analysis_status
        final_method = analysis_method
        final_norm = analysis_match_norm
        final_original = analysis_match_original
        final_line = analysis_line
        final_score = analysis_score
        final_score_composite = analysis_score_composite
        final_gap = analysis_gap
        final_flags = [flag for flag in analysis_flags.split("; ") if flag]

        if manual_status == "ACCEPT" and manual_norm:
            final_status = "ACEITO"
            final_method = "MANUAL_ACCEPT"
            final_norm = manual_norm
            final_original = manual_original or original_map.get(manual_norm, "")
            final_line = manual_line if pd.notna(manual_line) else line_map.get(manual_norm, pd.NA)
            final_score = manual_score if pd.notna(manual_score) else final_score
            final_score_composite = manual_score if pd.notna(manual_score) else final_score_composite
            add_flag(final_flags, "MANUAL_OVERRIDE", True)
        elif manual_status == "NO_MATCH":
            final_status = "SEM_MATCH"
            final_method = "MANUAL_NO_MATCH"
            final_norm = ""
            final_original = ""
            final_line = pd.NA
            final_score = pd.NA
            final_score_composite = pd.NA
            final_gap = pd.NA
            add_flag(final_flags, "MANUAL_OVERRIDE", True)
        elif manual_status == "REVIEW":
            final_status = "REVISAR"
            final_method = "MANUAL_REVIEW"
            add_flag(final_flags, "MANUAL_OVERRIDE", True)

        if manual_note:
            add_flag(final_flags, "MANUAL_NOTE", True)

        provisional.append(
            {
                "row_index": row_index,
                "final_status": final_status,
                "final_method": final_method,
                "final_match_t2_norm": final_norm,
                "final_match_t2_original": final_original,
                "final_line_match_t2": final_line,
                "final_score": final_score,
                "final_score_composite": final_score_composite,
                "final_score_gap": final_gap,
                "final_flags": final_flags,
                "manual_status": manual_status,
                "manual_sequence": results_df.at[row_index, "manual_sequence"],
                "final_match2_t1": analysis_match2_t1,
                "final_match2_t2": analysis_match2_t2,
                "final_match2_equal_prefix": analysis_match2_equal_prefix,
                "final_match2_score": analysis_match2_score,
            }
        )

    provisional_df = pd.DataFrame(provisional)
    accepted_df = provisional_df[
        (provisional_df["final_status"] == "ACEITO") & (provisional_df["final_match_t2_norm"] != "")
    ].copy()

    keep_rows: set[int] = set()
    if not accepted_df.empty:
        accepted_df["manual_priority"] = accepted_df["manual_status"].eq("ACCEPT").astype(int)
        accepted_df["score_priority"] = accepted_df["final_score"].apply(lambda value: safe_float(value) or 0.0)
        accepted_df["sequence_priority"] = accepted_df["manual_sequence"].apply(lambda value: safe_float(value) or 0.0)
        accepted_df = accepted_df.sort_values(
            ["final_match_t2_norm", "manual_priority", "score_priority", "sequence_priority", "row_index"],
            ascending=[True, False, False, True, True],
        )
        accepted_df["quota_order"] = accepted_df.groupby("final_match_t2_norm").cumcount() + 1
        accepted_df["quota_limit"] = accepted_df["final_match_t2_norm"].map(quota_map).fillna(0).astype(int)
        keep_rows = set(accepted_df.loc[accepted_df["quota_order"] <= accepted_df["quota_limit"], "row_index"].tolist())

    for row in provisional:
        row_index = row["row_index"]
        final_status = row["final_status"]
        final_method = row["final_method"]
        final_norm = row["final_match_t2_norm"]
        final_original = row["final_match_t2_original"]
        final_flags = row["final_flags"]
        quota_limit = quota_map.get(final_norm) if final_norm else None
        quota_order = pd.NA
        within_quota: Any = pd.NA
        source_norm = str(results_df.at[row_index, "nome_t1_match_norm"] or results_df.at[row_index, "nome_t1_norm"] or "")
        group_norm = final_norm
        group_original = final_original

        if final_status == "ACEITO" and final_norm:
            matching_rows = accepted_df[accepted_df["row_index"] == row_index]
            if not matching_rows.empty:
                quota_order = int(matching_rows.iloc[0]["quota_order"])
                quota_limit = int(matching_rows.iloc[0]["quota_limit"])
                within_quota = quota_order <= quota_limit
            else:
                within_quota = False

            if row_index not in keep_rows:
                if config and config.get("quantity_resolution_mode") == "marcar_excedente_excel1" and source_norm == final_norm:
                    final_status = "SEM_MATCH"
                    final_method = "EXCESS_LEFT"
                    add_flag(final_flags, "EXCESS_LEFT", True)
                else:
                    final_status = "REVISAR"
                    final_method = "FINAL_QUOTA_CONFLICT"
                    add_flag(final_flags, "FINAL_QUOTA_CONFLICT", True)
        elif (
            config
            and config.get("quantity_resolution_mode") == "marcar_excedente_excel1"
            and final_status == "REVISAR"
            and final_norm
            and source_norm == final_norm
            and final_method in {"QUOTA_CONFLICT", "FINAL_QUOTA_CONFLICT"}
        ):
            final_status = "SEM_MATCH"
            final_method = "EXCESS_LEFT"
            add_flag(final_flags, "EXCESS_LEFT", True)

        results_df.at[row_index, "final_status"] = final_status
        results_df.at[row_index, "final_method"] = final_method
        results_df.at[row_index, "final_group_t2_norm"] = group_norm if group_norm else ""
        results_df.at[row_index, "final_group_t2_original"] = group_original if group_original else ""
        results_df.at[row_index, "final_match_t2_norm"] = row["final_match_t2_norm"] if final_status == "ACEITO" else (
            row["final_match_t2_norm"] if final_status == "REVISAR" else ""
        )
        results_df.at[row_index, "final_match_t2_original"] = row["final_match_t2_original"] if final_status != "SEM_MATCH" else ""
        results_df.at[row_index, "final_line_match_t2"] = row["final_line_match_t2"] if final_status != "SEM_MATCH" else pd.NA
        results_df.at[row_index, "final_score"] = row["final_score"] if final_status != "SEM_MATCH" else pd.NA
        results_df.at[row_index, "final_score_composite"] = row["final_score_composite"] if final_status != "SEM_MATCH" else pd.NA
        results_df.at[row_index, "final_score_gap"] = row["final_score_gap"] if final_status != "SEM_MATCH" else pd.NA
        results_df.at[row_index, "final_match2_t1"] = row.get("final_match2_t1", "")
        results_df.at[row_index, "final_match2_t2"] = row.get("final_match2_t2", "")
        results_df.at[row_index, "final_match2_equal_prefix"] = bool(row.get("final_match2_equal_prefix", False))
        results_df.at[row_index, "final_match2_score"] = row.get("final_match2_score", pd.NA)
        final_norm_for_metrics = str(results_df.at[row_index, "final_match_t2_norm"] or group_norm or "")
        if (final_status != "SEM_MATCH" or final_method == "EXCESS_LEFT") and final_norm_for_metrics:
            ordered_score = ordered_character_ratio(source_norm, final_norm_for_metrics)
            aligned_score = aligned_character_ratio(source_norm, final_norm_for_metrics)
            length_gap = abs(len(source_norm) - len(final_norm_for_metrics))
            same_name_length = len(source_norm) == len(final_norm_for_metrics)
        else:
            ordered_score = 0.0
            aligned_score = 0.0
            length_gap = pd.NA
            same_name_length = pd.NA
        results_df.at[row_index, "final_score_ordered_chars"] = ordered_score if final_status != "SEM_MATCH" else pd.NA
        results_df.at[row_index, "final_score_aligned_chars"] = aligned_score if final_status != "SEM_MATCH" else pd.NA
        results_df.at[row_index, "final_name_length_gap"] = length_gap
        results_df.at[row_index, "final_same_name_length"] = same_name_length
        results_df.at[row_index, "final_color_bucket"] = determine_color_bucket(
            final_status,
            row["final_score"] if final_status != "SEM_MATCH" else pd.NA,
            ordered_score,
            aligned_score,
            bool(same_name_length) if pd.notna(same_name_length) else False,
            final_method=final_method,
            source_norm=source_norm,
            final_norm=final_norm_for_metrics,
            match2_equal_prefix=bool(row.get("final_match2_equal_prefix", False)),
            match2_active=bool(config and config.get("match2_col_t1") and config.get("match2_col_t2")),
            match2_values_present=bool(str(row.get("final_match2_t1", "") or "").strip() and str(row.get("final_match2_t2", "") or "").strip()),
            config=config,
        )
        results_df.at[row_index, "final_conflict_flags"] = flags_to_text(final_flags)
        results_df.at[row_index, "final_review_required"] = final_status == "REVISAR"
        results_df.at[row_index, "final_quota_limit"] = quota_limit if quota_limit is not None else pd.NA
        results_df.at[row_index, "final_quota_order"] = quota_order
        results_df.at[row_index, "final_within_quota"] = within_quota


def export_analysis_result(
    result: AnalysisResult,
    output_file: str | Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> Path:
    config = dict(result.config)
    if output_file is not None:
        config["output_file"] = str(output_file)
    output_path = Path(config["output_file"])

    emit_progress(progress_callback, "Atualizando estado final antes da exportação...", 10)
    recompute_final_state(result.results_df, result.catalog_df, config=result.config)
    result.summary_df = build_summary(result.results_df)
    result.review_df = result.results_df[result.results_df["final_status"] == "REVISAR"].copy()
    result.preview_df = result.results_df.head(40).copy()
    result.quota_df = build_quota_summary(result.results_df, result.catalog_df)

    source_original_columns = extract_original_columns(result.source_df, "source_row_id")
    target_original_columns = extract_original_columns(result.target_df, "target_row_id")
    export_source_df = result.source_df[source_original_columns].copy()
    export_target_df = result.target_df[target_original_columns].copy()
    results_export_df = build_results_export_df(result)
    summary_rows_df = build_export_summary_rows(result)
    reconciliation_df = build_grouped_reconciliation_df(result)
    result.grouped_reconciliation_df = reconciliation_df.copy()
    results_startrow = len(summary_rows_df) + 2

    emit_progress(progress_callback, "Escrevendo arquivo Excel...", 55)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_source_df.to_excel(writer, sheet_name="excel_1_original", index=False)
        export_target_df.to_excel(writer, sheet_name="excel_2_original", index=False)
        summary_rows_df.to_excel(writer, sheet_name="resultados_match", index=False)
        results_export_df.to_excel(writer, sheet_name="resultados_match", index=False, startrow=results_startrow)
        visible_cols = [column for column in reconciliation_df.columns if not column.startswith("_")]
        reconciliation_df[visible_cols].to_excel(writer, sheet_name="conciliacao_quantidades", index=False)

    emit_progress(progress_callback, "Formatando arquivo...", 88)
    format_output_workbook(output_path, result, results_startrow)
    emit_progress(progress_callback, f"Exportação concluída: {output_path}", 100)
    return output_path


def build_export_catalog(result: AnalysisResult) -> pd.DataFrame:
    """Return a safe copy of the export catalog.

    This helper is kept as a supported programmatic entry point for
    external scripts/integrations. It is intentionally retained even if
    in-repo callers are not present.
    """
    if result.catalog_df.empty:
        columns = ["nome_t2_norm", "nome_t2_original", "quota_original", "quota_limit", "excel_row_t2"]
        return pd.DataFrame(columns=columns)
    return result.catalog_df.copy()


def natural_sort_key(value: str) -> list[Any]:
    text = str(value)
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", text)]


def pick_primary_date_column(df: pd.DataFrame, columns: list[str]) -> str | None:
    best_column = None
    best_score = -1.0
    for column in columns:
        if column not in df.columns:
            continue
        parsed = pd.to_datetime(df[column], errors="coerce", dayfirst=True, format="mixed")
        valid_ratio = float(parsed.notna().mean()) if len(parsed) else 0.0
        if valid_ratio < 0.4:
            continue
        name = normalize_name(column)
        score = valid_ratio
        if "ATEND" in name or "DATA" in name or "DATE" in name or "LAUDO" in name or "ESTUDO" in name:
            score += 1.5
        if "NASC" in name or "BIRTH" in name:
            score -= 1.0
        if score > best_score:
            best_score = score
            best_column = column
    return best_column


def build_ordered_export_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    source_columns: list[str] = []
    for column in df.columns:
        if column == "source_row_id":
            break
        source_columns.append(column)

    primary_columns = [column for column in ("nome_t1_original", "final_match_t2_original") if column in df.columns]
    remaining_source = [column for column in source_columns if column not in primary_columns]
    helper_columns = [column for column in df.columns if column not in primary_columns and column not in remaining_source]

    ordered_columns = (
        primary_columns
        + sorted(remaining_source, key=natural_sort_key)
        + sorted(helper_columns, key=natural_sort_key)
    )
    ordered_df = df.loc[:, ordered_columns].copy()

    date_column = pick_primary_date_column(ordered_df, remaining_source)
    if date_column:
        date_values = pd.to_datetime(ordered_df[date_column], errors="coerce", dayfirst=True, format="mixed")
        ordered_df = (
            ordered_df.assign(_sort_date=date_values)
            .sort_values("_sort_date", ascending=False, kind="stable", na_position="last")
            .drop(columns="_sort_date")
        )
    return ordered_df.reset_index(drop=True)


def extract_original_columns(df: pd.DataFrame, marker_column: str) -> list[str]:
    columns: list[str] = []
    for column in df.columns:
        if column == marker_column:
            break
        columns.append(column)
    return columns


def build_results_export_df(result: AnalysisResult) -> pd.DataFrame:
    mode = str(result.config.get("output_mode", "enxuta")).strip().lower()
    if mode == "tecnica":
        return build_ordered_export_df(result.results_df)

    columns = [
        "excel_row_t1",
        "nome_t1_original",
        "final_status",
        "final_method",
        "final_match_t2_original",
        "final_score",
        "final_score_composite",
        "final_match2_t1",
        "final_match2_t2",
        "final_match2_equal_prefix",
        "final_match2_score",
        "final_color_bucket",
        "final_conflict_flags",
    ]
    return result.results_df[[column for column in columns if column in result.results_df.columns]].copy()


def build_export_summary_rows(result: AnalysisResult) -> pd.DataFrame:
    results_df = result.results_df
    source_total = len(result.source_df)
    target_total = len(result.target_df)
    accepted = int((results_df["final_status"] == "ACEITO").sum())
    review = int((results_df["final_status"] == "REVISAR").sum())
    no_match = int((results_df["final_status"] == "SEM_MATCH").sum())
    excess_left = int((results_df["final_method"] == "EXCESS_LEFT").sum())
    used_targets = int(results_df.loc[results_df["final_match_t2_norm"] != "", "final_group_t2_norm"].replace("", pd.NA).dropna().shape[0])
    unused_targets = max(target_total - int((result.results_df["final_status"] == "ACEITO").sum()), 0)
    bucket_counts = results_df["final_color_bucket"].fillna("SEM_BUCKET").value_counts().to_dict()
    lines = [
        ("Métrica", "Valor"),
        ("Total de linhas Excel 1", source_total),
        ("Total de linhas Excel 2", target_total),
        ("Diferença total (Excel 1 - Excel 2)", source_total - target_total),
        ("Aceitos no Excel 1", accepted),
        ("Revisão no Excel 1", review),
        ("Sem match no Excel 1", no_match),
        ("Excedente por quantidade no Excel 1", excess_left),
        ("Linhas utilizadas do Excel 2", target_total - unused_targets),
        ("Linhas não utilizadas do Excel 2", unused_targets),
    ]
    for bucket, count in sorted(bucket_counts.items()):
        lines.append((f"Bucket {bucket}", count))
    return pd.DataFrame(lines, columns=["Métrica", "Valor"])


def _parse_optional_datetime(value: Any) -> pd.Timestamp | None:
    text = str(value or "").strip()
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True, format="mixed")
    if pd.isna(parsed):
        return None
    return parsed


def _pair_group_rows_by_match2(
    left_rows: list[dict[str, Any]],
    right_rows: list[dict[str, Any]],
    match2_prefix_chars: int,
) -> list[tuple[dict[str, Any] | None, dict[str, Any] | None]]:
    if not left_rows and not right_rows:
        return []
    if not left_rows:
        return [(None, right) for right in right_rows]
    if not right_rows:
        return [(left, None) for left in left_rows]

    prefix_chars = max(int(match2_prefix_chars or 1), 1)
    remaining_right = set(range(len(right_rows)))
    assigned_right_by_left: dict[int, int] = {}

    def normalize_prefix(value: Any) -> str:
        return str(value or "").strip().upper()[:prefix_chars]

    # 1) Exact/Prefix-equal first when match2 exists on both sides.
    for left_idx, left in enumerate(left_rows):
        left_prefix = normalize_prefix(left.get("final_match2_t1", ""))
        if not left_prefix:
            continue
        exact_candidates = [
            right_idx
            for right_idx in remaining_right
            if normalize_prefix(right_rows[right_idx].get("match2_t2_original", "")) == left_prefix
        ]
        if exact_candidates:
            chosen = min(exact_candidates, key=lambda idx: int(right_rows[idx].get("target_row_id", 0)))
            assigned_right_by_left[left_idx] = chosen
            remaining_right.remove(chosen)

    # 2) If no exact match, align by date proximity when parseable.
    for left_idx, left in enumerate(left_rows):
        if left_idx in assigned_right_by_left:
            continue
        left_dt = _parse_optional_datetime(left.get("final_match2_t1", ""))
        if left_dt is None:
            continue
        dated_candidates: list[tuple[int, float]] = []
        for right_idx in remaining_right:
            right_dt = _parse_optional_datetime(right_rows[right_idx].get("match2_t2_original", ""))
            if right_dt is None:
                continue
            delta = abs((left_dt - right_dt).total_seconds())
            dated_candidates.append((right_idx, delta))
        if dated_candidates:
            chosen = min(dated_candidates, key=lambda item: (item[1], int(right_rows[item[0]].get("target_row_id", 0))))[0]
            assigned_right_by_left[left_idx] = chosen
            remaining_right.remove(chosen)

    # 3) Fallback: keep deterministic order for leftovers.
    for left_idx, _left in enumerate(left_rows):
        if left_idx in assigned_right_by_left:
            continue
        if not remaining_right:
            break
        chosen = min(remaining_right, key=lambda idx: int(right_rows[idx].get("target_row_id", 0)))
        assigned_right_by_left[left_idx] = chosen
        remaining_right.remove(chosen)

    pairs: list[tuple[dict[str, Any] | None, dict[str, Any] | None]] = []
    for left_idx, left in enumerate(left_rows):
        right = right_rows[assigned_right_by_left[left_idx]] if left_idx in assigned_right_by_left else None
        pairs.append((left, right))
    for right_idx in sorted(remaining_right, key=lambda idx: int(right_rows[idx].get("target_row_id", 0))):
        pairs.append((None, right_rows[right_idx]))
    return pairs


def build_grouped_reconciliation_df(result: AnalysisResult) -> pd.DataFrame:
    results_df = result.results_df.copy()
    target_df = result.target_df.copy()
    if results_df.empty and target_df.empty:
        return pd.DataFrame(columns=["Excel 1", "Excel 2", "_bucket_left", "_bucket_right"])

    target_records = target_df.to_dict("records")
    targets_by_norm: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in target_records:
        targets_by_norm[str(record.get("nome_t2_norm") or "")].append(record)

    for norm in targets_by_norm:
        targets_by_norm[norm].sort(key=lambda item: int(item.get("target_row_id", 0)))

    result_records = results_df.to_dict("records")
    extra_cols_t1 = parse_csv_columns(result.config.get("tab4_extra_cols_t1", ""))
    extra_cols_t2 = parse_csv_columns(result.config.get("tab4_extra_cols_t2", ""))
    money_cols_t1 = set(parse_csv_columns(result.config.get("tab4_money_cols_t1", "")))
    money_cols_t2 = set(parse_csv_columns(result.config.get("tab4_money_cols_t2", "")))
    if result.config.get("match2_col_t1"):
        source_match2_col_index = excel_col_to_index(result.config["match2_col_t1"])
        source_original = extract_original_columns(result.source_df, "source_row_id")
        if source_match2_col_index < len(source_original):
            default_col = source_original[source_match2_col_index]
            if default_col not in extra_cols_t1:
                extra_cols_t1.insert(0, default_col)
    if result.config.get("match2_col_t2"):
        target_match2_col_index = excel_col_to_index(result.config["match2_col_t2"])
        target_original = extract_original_columns(result.target_df, "target_row_id")
        if target_match2_col_index < len(target_original):
            default_col = target_original[target_match2_col_index]
            if default_col not in extra_cols_t2:
                extra_cols_t2.insert(0, default_col)

    source_records_by_id = {
        int(row["source_row_id"]): row for row in result.source_df.to_dict("records")
    } if not result.source_df.empty else {}
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in result_records:
        group_key = str(record.get("final_group_t2_norm") or record.get("nome_t1_norm") or "")
        groups[group_key].append(record)

    all_group_keys = sorted(set(groups.keys()) | set(targets_by_norm.keys()))
    rows: list[dict[str, Any]] = []
    for group_key in all_group_keys:
        left_rows = groups.get(group_key, [])
        left_rows.sort(
            key=lambda item: (
                {"EXACT": 0, "MATCH": 1, "REVIEW": 2, "EXCESS_LEFT": 3, "NO_MATCH": 4}.get(
                    str(item.get("final_color_bucket") or "NO_MATCH"), 9
                ),
                0 if bool(item.get("final_match2_equal_prefix", False)) else 1,
                -(safe_float(item.get("final_match2_score")) or 0.0),
                int(item.get("source_row_id", 0)),
            )
        )
        right_rows = targets_by_norm.get(group_key, [])
        pairs = _pair_group_rows_by_match2(
            left_rows,
            right_rows,
            int(result.config.get("match2_prefix_chars", 8)),
        )
        if not pairs:
            pairs = [(None, None)]
        for left, right in pairs:
            left_bucket = str(left.get("final_color_bucket") or "NO_MATCH") if left else ""
            right_bucket = left_bucket if right and left else ("NO_MATCH" if right else "")
            rows.append(
                {
                    "Excel 1": str(left.get("nome_t1_original") or "") if left else "",
                    "Excel 2": str(right.get("nome_t2_original") or "") if right else "",
                    "_bucket_left": left_bucket,
                    "_bucket_right": right_bucket,
                    "_group_key": group_key,
                }
            )
            if left:
                source_info = source_records_by_id.get(int(left.get("source_row_id", 0)), {})
                for column in extra_cols_t1:
                    raw_value = source_info.get(column, "")
                    if column in money_cols_t1:
                        numeric_value = parse_brl_currency_value(raw_value)
                        rows[-1][f"E1:{column}"] = numeric_value if numeric_value is not None else str(raw_value or "")
                    else:
                        rows[-1][f"E1:{column}"] = str(raw_value or "")
            else:
                for column in extra_cols_t1:
                    rows[-1][f"E1:{column}"] = ""
            if right:
                for column in extra_cols_t2:
                    raw_value = right.get(column, "")
                    if column in money_cols_t2:
                        numeric_value = parse_brl_currency_value(raw_value)
                        rows[-1][f"E2:{column}"] = numeric_value if numeric_value is not None else str(raw_value or "")
                    else:
                        rows[-1][f"E2:{column}"] = str(raw_value or "")
            else:
                for column in extra_cols_t2:
                    rows[-1][f"E2:{column}"] = ""
        rows.append({"Excel 1": "", "Excel 2": "", "_bucket_left": "", "_bucket_right": "", "_group_key": ""})
    return pd.DataFrame(rows)


def run_matching(config: dict[str, Any], progress_callback: ProgressCallback | None = None) -> Path:
    """Run full analysis + export pipeline.

    This helper is kept as a supported programmatic entry point for
    external scripts/integrations. It is intentionally retained even if
    in-repo callers are not present.
    """
    ensure_supported_python_version()
    result = analyze_matching(config, progress_callback=progress_callback)
    return export_analysis_result(result, progress_callback=progress_callback)


# =========================
# GUI
# =========================


class ToolTip:
    def __init__(self, widget, text: str) -> None:
        self.widget = widget
        self.text = text
        self.tip_window: tk.Toplevel | None = None
        self.widget.bind("<Enter>", self.show)
        self.widget.bind("<Leave>", self.hide)

    def show(self, _event=None) -> None:
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 16
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            self.tip_window,
            text=self.text,
            justify="left",
            background="#FFF9D6",
            relief="solid",
            borderwidth=1,
            padx=8,
            pady=6,
            wraplength=280,
        )
        label.pack()

    def hide(self, _event=None) -> None:
        if self.tip_window is not None:
            self.tip_window.destroy()
            self.tip_window = None


class MatcherApp:
    def __init__(self, root: Any):
        self.root = root
        self.root.title(f"Matcher de Nomes - Análise / Revisão / Exportação {APP_VERSION}")
        self.root.geometry("1480x980")
        self.root.minsize(1260, 860)
        self.last_output_path: Path | None = None
        self.analysis_result: AnalysisResult | None = None
        self.catalog_df: pd.DataFrame = pd.DataFrame()
        self.manual_sequence = 0
        self.last_manual_actions: list[str] = []
        self.busy_operation: str | None = None
        self._last_workbook_metadata_error: str = ""

        self.vars: dict[str, tk.Variable] = {
            "input_file_t1": tk.StringVar(),
            "input_file_t2": tk.StringVar(),
            "output_file": tk.StringVar(value="resultado_matching.xlsx"),
            "sheet_t1": tk.StringVar(),
            "sheet_t2": tk.StringVar(),
            "header_row_t1": tk.StringVar(value="1"),
            "header_row_t2": tk.StringVar(value="1"),
            "name_col_t1": tk.StringVar(value="C"),
            "name_col_t2": tk.StringVar(value="E"),
            "match2_col_t1": tk.StringVar(value=""),
            "match2_col_t2": tk.StringVar(value=""),
            "match2_prefix_chars": tk.StringVar(value="8"),
            "match2_weight": tk.StringVar(value="0.2"),
            "tab4_extra_cols_t1": tk.StringVar(value=""),
            "tab4_extra_cols_t2": tk.StringVar(value=""),
            "tab4_money_cols_t1": tk.StringVar(value=""),
            "tab4_money_cols_t2": tk.StringVar(value=""),
            "output_mode": tk.StringVar(value="enxuta"),
            "quantity_resolution_mode": tk.StringVar(value="marcar_excedente_excel1"),
            "color_exact": tk.StringVar(value=DEFAULT_MATCH_COLOR_LABELS["color_exact"]),
            "color_match": tk.StringVar(value=DEFAULT_MATCH_COLOR_LABELS["color_match"]),
            "color_review": tk.StringVar(value=DEFAULT_MATCH_COLOR_LABELS["color_review"]),
            "color_no_match": tk.StringVar(value=DEFAULT_MATCH_COLOR_LABELS["color_no_match"]),
            "color_excess_left": tk.StringVar(value=DEFAULT_MATCH_COLOR_LABELS["color_excess_left"]),
            "max_external_chars": tk.StringVar(value="30"),
            "accept_score": tk.StringVar(value="92"),
            "review_score": tk.StringVar(value="85"),
            "min_gap_for_accept": tk.StringVar(value="4"),
            "top_candidates_to_keep": tk.StringVar(value="5"),
            "allow_reuse_t2_matches": tk.BooleanVar(value=False),
            "max_matches_per_t2_name": tk.StringVar(value="3"),
            "auto_open_output": tk.BooleanVar(value=True),
            "weight_token_set": tk.StringVar(value="27"),
            "weight_partial": tk.StringVar(value="21"),
            "weight_sort": tk.StringVar(value="15"),
            "weight_prefix": tk.StringVar(value="15"),
            "weight_ordered_chars": tk.StringVar(value="14"),
            "weight_aligned_chars": tk.StringVar(value="8"),
            "length_gap_penalty_per_char": tk.StringVar(value="0.5"),
            "max_length_gap_penalty": tk.StringVar(value="10"),
            "missing_surname_penalty": tk.StringVar(value="3"),
        }
        self.status_var = tk.StringVar(value="Pronto.")
        self.progress_var = tk.DoubleVar(value=0.0)
        self.manual_note_var = tk.StringVar()
        self.quick_preset_var = tk.StringVar(value="Equilibrado")
        self.config_mode_var = tk.StringVar(value="Básico")
        self.review_filter_var = tk.StringVar(value="Todos")
        self.review_search_var = tk.StringVar()
        self.review_hint_var = tk.StringVar(value="Selecione uma linha da fila para ver motivos, candidatos e ações rápidas.")
        self.export_snapshot_var = tk.StringVar(value="Nenhuma exportação executada nesta sessão.")
        self.match1_limit_hint_var = tk.StringVar(value="Match 1 efetivo: 30 caracteres")
        self.sheet_options_t1: list[str] = []
        self.sheet_options_t2: list[str] = []
        self.column_options_t1: list[str] = []
        self.column_options_t2: list[str] = []
        self.summary_card_vars = {
            "total": tk.StringVar(value="0"),
            "accepted": tk.StringVar(value="0"),
            "review": tk.StringVar(value="0"),
            "no_match": tk.StringVar(value="0"),
        }
        self.style = ttk.Style()
        self._configure_styles()

        self._build_ui()
        self._bind_shortcuts()
        self.vars["max_external_chars"].trace_add("write", lambda *_args: self.update_match1_limit_hint())
        self.load_ui_state()
        self.update_match1_limit_hint()
        self.refresh_sheet_choices()
        self.update_config_mode()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self) -> None:
        container = ttk.Frame(self.root, padding=10)
        container.pack(fill="both", expand=True)

        header = ttk.Frame(container)
        header.pack(fill="x", pady=(0, 10))
        title_row = ttk.Frame(header)
        title_row.pack(fill="x")
        title_text = "Matcher V2"
        if HAS_TTKBOOTSTRAP:
            title_text += " • ttkbootstrap"
        ttk.Label(title_row, text=title_text, font=("Segoe UI", 16, "bold")).pack(side="left", anchor="w")
        ttk.Button(title_row, text="Ajuda rápida", command=self.show_help, style="Info.TButton").pack(side="right", padx=(8, 0))
        ttk.Button(title_row, text="Sobre", command=self.show_about, style="Primary.TButton").pack(side="right")
        ttk.Label(
            header,
            text="Validação -> Análise -> Revisão Manual -> Exportação",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(2, 0))

        command_row = ttk.LabelFrame(container, text="Ações rápidas", padding=10)
        command_row.pack(fill="x", pady=(0, 10))
        ttk.Label(command_row, text="Preset").pack(side="left")
        preset_box = ttk.Combobox(
            command_row,
            textvariable=self.quick_preset_var,
            values=list(QUICK_PRESETS.keys()),
            state="readonly",
            width=18,
        )
        preset_box.pack(side="left", padx=(8, 8))
        ttk.Button(command_row, text="Aplicar preset", command=self.apply_quick_preset, style="Info.TButton").pack(side="left")
        ttk.Button(command_row, text="Validar", command=self.validate_and_preview, style="Warning.TButton").pack(side="left", padx=(8, 0))
        ttk.Button(command_row, text="Executar análise", command=self.start_analysis, style="Primary.TButton").pack(side="left", padx=(8, 0))
        ttk.Button(command_row, text="Exportar", command=self.start_export, style="Success.TButton").pack(side="left", padx=(8, 0))
        ttk.Label(command_row, text="Atalhos: Ctrl+O abrir | F5 validar | Ctrl+R analisar | Ctrl+E exportar").pack(
            side="right"
        )

        status_bar = ttk.Frame(container)
        status_bar.pack(fill="x", pady=(0, 10))
        ttk.Label(status_bar, textvariable=self.status_var, font=("Segoe UI", 10, "bold")).pack(side="left")
        ttk.Progressbar(status_bar, variable=self.progress_var, maximum=100, length=360).pack(side="right")

        self.notebook = ttk.Notebook(container)
        self.notebook.pack(fill="both", expand=True)

        self.tab_config = ttk.Frame(self.notebook, padding=10)
        self.tab_analyze = ttk.Frame(self.notebook, padding=10)
        self.tab_review = ttk.Frame(self.notebook, padding=10)
        self.tab_export = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.tab_config, text="Configuração")
        self.notebook.add(self.tab_analyze, text="Análise")
        self.notebook.add(self.tab_review, text="Revisão")
        self.notebook.add(self.tab_export, text="Exportação")

        self._build_config_scroll_container()
        self._build_config_tab()
        self._build_analyze_tab()
        self._build_review_tab()
        self._build_export_tab()

        self.log("Aplicação iniciada.")

    def _build_config_scroll_container(self) -> None:
        self.tab_config_canvas = tk.Canvas(
            self.tab_config,
            highlightthickness=0,
            bg=UI_COLORS["bg"],
        )
        self.tab_config_scrollbar = ttk.Scrollbar(
            self.tab_config,
            orient="vertical",
            command=self.tab_config_canvas.yview,
        )
        self.tab_config_canvas.configure(yscrollcommand=self.tab_config_scrollbar.set)
        self.tab_config_scrollbar.pack(side="right", fill="y")
        self.tab_config_canvas.pack(side="left", fill="both", expand=True)
        self.tab_config_content = ttk.Frame(self.tab_config_canvas, padding=2)
        self._tab_config_window_id = self.tab_config_canvas.create_window(
            (0, 0),
            window=self.tab_config_content,
            anchor="nw",
        )
        self.tab_config_content.bind(
            "<Configure>",
            lambda _event: self.tab_config_canvas.configure(scrollregion=self.tab_config_canvas.bbox("all")),
        )
        self.tab_config_canvas.bind(
            "<Configure>",
            lambda event: self.tab_config_canvas.itemconfigure(self._tab_config_window_id, width=event.width),
        )
        self.tab_config_canvas.bind("<Enter>", self._bind_config_mousewheel)
        self.tab_config_canvas.bind("<Leave>", self._unbind_config_mousewheel)

    def _bind_config_mousewheel(self, _event=None) -> None:
        self.tab_config_canvas.bind_all("<MouseWheel>", self._on_config_mousewheel)

    def _unbind_config_mousewheel(self, _event=None) -> None:
        self.tab_config_canvas.unbind_all("<MouseWheel>")

    def _on_config_mousewheel(self, event) -> None:
        self.tab_config_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def update_match1_limit_hint(self) -> None:
        raw = str(self.vars["max_external_chars"].get() or "").strip()
        try:
            value = int(raw)
            if value <= 0:
                raise ValueError()
            self.match1_limit_hint_var.set(f"Match 1 efetivo: {value} caracteres")
        except ValueError:
            self.match1_limit_hint_var.set("Match 1 efetivo: valor inválido (use inteiro > 0)")

    def _configure_styles(self) -> None:
        try:
            self.root.configure(background=UI_COLORS["bg"])
            self.style.configure(".", background=UI_COLORS["bg"], foreground=UI_COLORS["text"])
            self.style.configure("TFrame", background=UI_COLORS["bg"])
            self.style.configure("TLabel", background=UI_COLORS["bg"], foreground=UI_COLORS["text"])
            self.style.configure("TCheckbutton", background=UI_COLORS["bg"], foreground=UI_COLORS["text"])
            self.style.configure(
                "TLabelframe",
                background=UI_COLORS["panel"],
                bordercolor=UI_COLORS["border"],
                relief="solid",
                padding=8,
            )
            self.style.configure(
                "TLabelframe.Label",
                background=UI_COLORS["panel"],
                foreground=UI_COLORS["accent"],
                font=("Segoe UI", 10, "bold"),
            )
            self.style.configure(
                "TNotebook",
                background=UI_COLORS["bg"],
                borderwidth=0,
            )
            self.style.configure(
                "TNotebook.Tab",
                padding=(14, 8),
                background=UI_COLORS["panel"],
                foreground=UI_COLORS["text"],
            )
            self.style.map(
                "TNotebook.Tab",
                background=[("selected", UI_COLORS["btn_primary"]), ("active", UI_COLORS["accent"])],
                foreground=[("selected", "#FFFFFF"), ("active", "#FFFFFF")],
            )
            self.style.configure(
                "TButton",
                padding=(10, 6),
                borderwidth=0,
                background=UI_COLORS["panel_alt"],
                foreground=UI_COLORS["text"],
            )
            self.style.map(
                "TButton",
                background=[("active", UI_COLORS["card"]), ("pressed", UI_COLORS["btn_primary"])],
                foreground=[("active", UI_COLORS["text"]), ("pressed", "#FFFFFF")],
            )
            self.style.configure("Primary.TButton", background=UI_COLORS["btn_primary"], foreground="#FFFFFF")
            self.style.map(
                "Primary.TButton",
                background=[("active", "#3F91F8"), ("pressed", "#2168C8")],
                foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
            )
            self.style.configure("Success.TButton", background=UI_COLORS["btn_success"], foreground="#FFFFFF")
            self.style.map(
                "Success.TButton",
                background=[("active", "#2BD97A"), ("pressed", "#149B54")],
                foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
            )
            self.style.configure("Warning.TButton", background=UI_COLORS["btn_warning"], foreground="#FFFFFF")
            self.style.map(
                "Warning.TButton",
                background=[("active", "#F6AD2C"), ("pressed", "#CF7E07")],
                foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
            )
            self.style.configure("Danger.TButton", background=UI_COLORS["btn_danger"], foreground="#FFFFFF")
            self.style.map(
                "Danger.TButton",
                background=[("active", "#EF5D4E"), ("pressed", "#C0392B")],
                foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
            )
            self.style.configure("Info.TButton", background=UI_COLORS["btn_info"], foreground="#FFFFFF")
            self.style.map(
                "Info.TButton",
                background=[("active", "#18BFE0"), ("pressed", "#0089A8")],
                foreground=[("active", "#FFFFFF"), ("pressed", "#FFFFFF")],
            )
            self.style.configure(
                "TEntry",
                fieldbackground=UI_COLORS["field"],
                background=UI_COLORS["field"],
                foreground=UI_COLORS["text"],
                bordercolor=UI_COLORS["border"],
                lightcolor=UI_COLORS["field"],
                darkcolor=UI_COLORS["field"],
                insertcolor=UI_COLORS["text"],
            )
            self.style.configure(
                "TCombobox",
                fieldbackground=UI_COLORS["field"],
                background=UI_COLORS["field"],
                foreground=UI_COLORS["text"],
                bordercolor=UI_COLORS["border"],
                arrowcolor=UI_COLORS["text"],
                lightcolor=UI_COLORS["field"],
                darkcolor=UI_COLORS["field"],
            )
            self.style.map(
                "TCombobox",
                fieldbackground=[("readonly", UI_COLORS["field"])],
                selectbackground=[("readonly", UI_COLORS["field"])],
                selectforeground=[("readonly", UI_COLORS["text"])],
            )
            self.style.configure("Treeview", rowheight=30, font=("Segoe UI", 10))
            self.style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
            self.style.configure(
                "Treeview",
                rowheight=30,
                font=("Segoe UI", 10),
                background=UI_COLORS["card"],
                fieldbackground=UI_COLORS["card"],
                foreground=UI_COLORS["text"],
                bordercolor=UI_COLORS["border"],
            )
            self.style.configure(
                "Treeview.Heading",
                font=("Segoe UI", 10, "bold"),
                background=UI_COLORS["btn_primary"],
                foreground="#FFFFFF",
            )
            self.style.map(
                "Treeview",
                background=[("selected", UI_COLORS["btn_info"])],
                foreground=[("selected", "#FFFFFF")],
            )
        except Exception:
            pass

    def _bind_shortcuts(self) -> None:
        self.root.bind_all("<Control-o>", lambda _event: self.pick_input_file_t1())
        self.root.bind_all("<F5>", lambda _event: self.validate_and_preview())
        self.root.bind_all("<Control-r>", lambda _event: self.start_analysis())
        self.root.bind_all("<Control-e>", lambda _event: self.start_export())
        self.root.bind_all("<Control-s>", lambda _event: self.save_ui_state(show_feedback=True))

    def _configure_status_tree(self, tree: Any) -> None:
        tree.tag_configure("ACEITO", background="#24563D", foreground="#EAFBF2")
        tree.tag_configure("REVISAR", background="#5A4B1E", foreground="#FFF5DA")
        tree.tag_configure("SEM_MATCH", background="#503339", foreground="#FFECEF")
        tree.tag_configure("CONFLITO", background="#2A4A63", foreground="#EAF6FF")

    def _draw_summary_chart(self, canvas: tk.Canvas, values: dict[str, int]) -> None:
        canvas.delete("all")
        width = max(canvas.winfo_width(), 420)
        height = max(canvas.winfo_height(), 120)
        canvas.configure(scrollregion=(0, 0, width, height), background=UI_COLORS["panel_alt"])
        total = max(sum(values.values()), 1)
        items = [
            ("ACEITO", values.get("ACEITO", 0), UI_COLORS["btn_success"]),
            ("REVISAR", values.get("REVISAR", 0), UI_COLORS["btn_warning"]),
            ("SEM_MATCH", values.get("SEM_MATCH", 0), UI_COLORS["btn_danger"]),
        ]
        start_x = 24
        start_y = 22
        bar_width = max(width - 220, 180)
        for index, (label, value, color) in enumerate(items):
            top = start_y + index * 30
            ratio = value / total
            canvas.create_text(
                start_x,
                top + 9,
                text=label,
                anchor="w",
                font=("Segoe UI", 9, "bold"),
                fill=UI_COLORS["text"],
            )
            canvas.create_rectangle(
                start_x + 100,
                top,
                start_x + 100 + bar_width,
                top + 18,
                fill="#405965",
                outline="",
            )
            canvas.create_rectangle(
                start_x + 100,
                top,
                start_x + 100 + max(int(bar_width * ratio), 2 if value else 0),
                top + 18,
                fill=color,
                outline="",
            )
            canvas.create_text(
                start_x + 100 + bar_width + 12,
                top + 9,
                text=f"{value} ({ratio * 100:.1f}%)",
                anchor="w",
                font=("Segoe UI", 9),
                fill=UI_COLORS["text"],
            )

    def _show_info(self, title: str, message: str) -> None:
        if HAS_TTKBOOTSTRAP and BootMessagebox is not None:
            try:
                BootMessagebox.show_info(message, title=title, parent=self.root)
                return
            except Exception:
                pass
        messagebox.showinfo(title, message)

    def _show_warning(self, title: str, message: str) -> None:
        if HAS_TTKBOOTSTRAP and BootMessagebox is not None:
            try:
                BootMessagebox.show_warning(message, title=title, parent=self.root)
                return
            except Exception:
                pass
        messagebox.showwarning(title, message)

    def _show_error(self, title: str, message: str) -> None:
        if HAS_TTKBOOTSTRAP and BootMessagebox is not None:
            try:
                BootMessagebox.show_error(message, title=title, parent=self.root)
                return
            except Exception:
                pass
        messagebox.showerror(title, message)

    def show_help(self) -> None:
        self._show_info(
            "Ajuda rápida",
            "Fluxo sugerido:\n\n"
            "1. Selecione a planilha de entrada.\n"
            "2. Valide o mapeamento.\n"
            "3. Execute a análise.\n"
            "4. Revise os casos pendentes.\n"
            "5. Exporte a planilha final.\n\n"
            "Atalhos:\n"
            "Ctrl+O abrir arquivo\n"
            "F5 validar\n"
            "Ctrl+R analisar\n"
            "Ctrl+E exportar\n"
            "Ctrl+S salvar configuração visual",
        )

    def show_about(self) -> None:
        toolkit_name = "ttkbootstrap" if HAS_TTKBOOTSTRAP else "ttk padrão"
        self._show_info(
            "Sobre o Matcher",
            f"Matcher de Nomes {APP_VERSION}\n\n"
            f"Interface atual: {toolkit_name}\n"
            "Objetivo: validar, analisar, revisar e exportar correspondências entre nomes de planilhas Excel.",
        )

    def apply_quick_preset(self) -> None:
        preset = QUICK_PRESETS.get(self.quick_preset_var.get())
        if not preset:
            return
        for key, value in preset.items():
            variable = self.vars.get(key)
            if variable is None:
                continue
            variable.set(value)
        self.log(f"Preset aplicado: {self.quick_preset_var.get()}")
        self.set_status(f"Preset '{self.quick_preset_var.get()}' aplicado.", None)

    def update_config_mode(self, *_args: Any) -> None:
        mode = self.config_mode_var.get()
        if hasattr(self, "advanced_frame"):
            if mode in {"Avançado", "Especialista"}:
                self.advanced_frame.pack(fill="x", pady=(0, 8))
            else:
                self.advanced_frame.pack_forget()
        if hasattr(self, "expert_frame"):
            if mode == "Especialista":
                self.expert_frame.pack(fill="x", pady=(0, 8))
            else:
                self.expert_frame.pack_forget()

    def save_ui_state(self, show_feedback: bool = False) -> None:
        payload = {
            "vars": {key: variable.get() for key, variable in self.vars.items()},
            "quick_preset": self.quick_preset_var.get(),
            "config_mode": self.config_mode_var.get(),
            "review_filter": self.review_filter_var.get(),
        }
        try:
            UI_STATE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            if show_feedback:
                self.set_status("Configuração visual salva.", None)
        except Exception as exc:
            if show_feedback:
                self._show_error("Salvar configuração", str(exc))

    def load_ui_state(self) -> None:
        if not UI_STATE_FILE.exists():
            return
        try:
            payload = json.loads(UI_STATE_FILE.read_text(encoding="utf-8"))
            legacy_input = payload.get("vars", {}).get("input_file")
            for key, value in payload.get("vars", {}).items():
                variable = self.vars.get(key)
                if variable is not None:
                    variable.set(value)
            if legacy_input and not self.vars["input_file_t1"].get():
                self.vars["input_file_t1"].set(legacy_input)
            for key in MATCH_COLOR_KEYS:
                self.normalize_color_var_value(key)
            self.quick_preset_var.set(payload.get("quick_preset", self.quick_preset_var.get()))
            self.config_mode_var.set(payload.get("config_mode", self.config_mode_var.get()))
            self.review_filter_var.set(payload.get("review_filter", self.review_filter_var.get()))
            self.log("Último estado visual carregado.")
        except Exception as exc:
            self.log(f"Não foi possível carregar o estado visual: {exc}")

    def on_close(self) -> None:
        self.save_ui_state()
        self.root.destroy()

    def _build_config_tab(self) -> None:
        config_parent = self.tab_config_content
        mode_frame = ttk.LabelFrame(config_parent, text="Modo de configuração", padding=8)
        mode_frame.pack(fill="x", pady=(0, 6))
        ttk.Label(mode_frame, text="Nível de detalhe").pack(side="left")
        mode_box = ttk.Combobox(
            mode_frame,
            textvariable=self.config_mode_var,
            values=["Básico", "Avançado", "Especialista"],
            state="readonly",
            width=18,
        )
        mode_box.pack(side="left", padx=(6, 10))
        mode_box.bind("<<ComboboxSelected>>", self.update_config_mode)
        ttk.Label(
            mode_frame,
            text="Básico mostra o essencial, Avançado libera cotas e Especialista expõe pesos finos.",
        ).pack(side="left")

        files_frame = ttk.LabelFrame(config_parent, text="Arquivos", padding=8)
        files_frame.pack(fill="x", pady=(0, 6))
        self._add_file_row(files_frame, "Planilha Excel 1", "input_file_t1", self.pick_input_file_t1, 0)
        self._add_file_row(files_frame, "Planilha Excel 2", "input_file_t2", self.pick_input_file_t2, 1)
        self._add_file_row(files_frame, "Planilha de saída", "output_file", self.pick_output_file, 2)

        workbook_frame = ttk.LabelFrame(config_parent, text="Mapeamento da planilha", padding=8)
        workbook_frame.pack(fill="x", pady=(0, 6))
        self._add_sheet_field(workbook_frame, 0, 0, "Aba Excel 1", "sheet_t1", "Aba principal do Excel 1; por padrão a primeira aba encontrada.")
        self._add_sheet_field(workbook_frame, 0, 3, "Aba Excel 2", "sheet_t2", "Aba principal do Excel 2; por padrão a primeira aba encontrada.")
        self._add_setting_field(workbook_frame, 1, 0, "Linha de cabeçalho Excel 1", "header_row_t1", "Número da linha (base 1) onde o cabeçalho começa no Excel 1.")
        self._add_setting_field(workbook_frame, 1, 3, "Linha de cabeçalho Excel 2", "header_row_t2", "Número da linha (base 1) onde o cabeçalho começa no Excel 2.")
        self._add_setting_field(workbook_frame, 2, 0, "Coluna de nome Excel 1", "name_col_t1", "Letra da coluna no Excel 1 que contém o nome a casar.")
        self._add_setting_field(workbook_frame, 2, 3, "Coluna de nome Excel 2", "name_col_t2", "Letra da coluna no Excel 2 que contém o nome alvo.")
        self._add_setting_field(workbook_frame, 3, 0, "Coluna match 2 Excel 1", "match2_col_t1", "Opcional: segunda coluna para comparação auxiliar (ex.: data).")
        self._add_setting_field(workbook_frame, 3, 3, "Coluna match 2 Excel 2", "match2_col_t2", "Opcional: segunda coluna para comparação auxiliar (ex.: data).")
        self._add_setting_field(workbook_frame, 4, 0, "Caracteres do match 2", "match2_prefix_chars", "Quantidade de caracteres usados na comparação auxiliar da coluna match 2.")
        self._add_setting_field(workbook_frame, 4, 3, "Peso do match 2", "match2_weight", "Peso auxiliar no score composto. Não bloqueia match por si só.")
        for column in (1, 4):
            workbook_frame.columnconfigure(column, weight=1)

        output_frame = ttk.LabelFrame(config_parent, text="Saída e cores", padding=8)
        output_frame.pack(fill="x", pady=(0, 6))
        self._add_combobox_field(
            output_frame,
            0,
            0,
            "Modo da aba 3",
            "output_mode",
            ["enxuta", "tecnica"],
            "Define se a terceira aba será enxuta ou técnica.",
        )
        self._add_combobox_field(
            output_frame,
            0,
            3,
            "Regra de quantidade",
            "quantity_resolution_mode",
            ["marcar_excedente_excel1"],
            "Quando houver mais ocorrências no Excel 1 que no Excel 2, apenas as disponíveis serão casadas e o restante ficará destacado.",
        )
        self._add_color_field(output_frame, 1, 0, "Cor match exato", "color_exact")
        self._add_color_field(output_frame, 1, 3, "Cor match aceito", "color_match")
        self._add_color_field(output_frame, 2, 0, "Cor revisão", "color_review")
        self._add_color_field(output_frame, 2, 3, "Cor sem match", "color_no_match")
        self._add_color_field(output_frame, 3, 0, "Cor excedente Excel 1", "color_excess_left")
        self._add_csv_column_field(
            output_frame,
            4,
            0,
            "Extras aba 4 (Excel 1)",
            "tab4_extra_cols_t1",
            "Colunas extras exibidas na aba 4 para Excel 1 (separadas por vírgula).",
            side="t1",
        )
        self._add_csv_column_field(
            output_frame,
            4,
            3,
            "Extras aba 4 (Excel 2)",
            "tab4_extra_cols_t2",
            "Colunas extras exibidas na aba 4 para Excel 2 (separadas por vírgula).",
            side="t2",
        )
        self._add_csv_column_field(
            output_frame,
            5,
            0,
            "Monetárias aba 4 (Excel 1)",
            "tab4_money_cols_t1",
            "Colunas da aba 4 do Excel 1 que devem ser exportadas como moeda (R$). Aceita valores simples e textos com R$. A coluna também precisa estar em Extras aba 4.",
            side="t1",
        )
        self._add_csv_column_field(
            output_frame,
            5,
            3,
            "Monetárias aba 4 (Excel 2)",
            "tab4_money_cols_t2",
            "Colunas da aba 4 do Excel 2 que devem ser exportadas como moeda (R$). Aceita valores simples e textos com R$. A coluna também precisa estar em Extras aba 4.",
            side="t2",
        )
        output_frame.columnconfigure(1, weight=1)
        output_frame.columnconfigure(4, weight=1)

        recommended_frame = ttk.LabelFrame(config_parent, text="Padrões recomendados", padding=8)
        recommended_frame.pack(fill="x", pady=(0, 6))
        ttk.Label(
            recommended_frame,
            textvariable=self.match1_limit_hint_var,
            font=("Segoe UI", 9, "bold"),
        ).grid(row=0, column=0, columnspan=6, sticky="w", padx=4, pady=(0, 4))
        recommended_fields = [
            ("Limite de caracteres (Match 1)", "max_external_chars", "Limita efetivamente os nomes do Match 1 aos primeiros N caracteres antes da comparação."),
            ("Pontuação de aceite", "accept_score", "Candidatos acima desta pontuação podem ser aceitos automaticamente quando estrutura e diferença também forem fortes."),
            ("Pontuação de revisão", "review_score", "Candidatos acima desta pontuação entram em revisão manual quando não forem seguros para aceite automático."),
            ("Diferença mínima", "min_gap_for_accept", "Diferença mínima para o próximo candidato necessária para aceite automático seguro."),
            ("Candidatos em prévia", "top_candidates_to_keep", "Quantidade de candidatos mantidos para pré-visualização e revisão manual."),
        ]
        for index, (label, key, tooltip) in enumerate(recommended_fields):
            row = index // 2 + 1
            col = (index % 2) * 3
            self._add_setting_field(recommended_frame, row, col, label, key, tooltip)
        for column in (1, 4):
            recommended_frame.columnconfigure(column, weight=1)

        self.advanced_frame = ttk.LabelFrame(config_parent, text="Controles avançados", padding=8)
        self._add_setting_field(
            self.advanced_frame,
            0,
            0,
            "Limite de reaproveitamento",
            "max_matches_per_t2_name",
            "Máximo de reaproveitamento por nome normalizado da T2 quando o reaproveitamento estiver ativo. A cota real continua sendo respeitada.",
        )
        self.advanced_frame.columnconfigure(1, weight=1)
        self.advanced_frame.columnconfigure(4, weight=1)
        ttk.Checkbutton(
            self.advanced_frame,
            text="Permitir reaproveitar nomes da T2 até o limite configurado",
            variable=self.vars["allow_reuse_t2_matches"],
        ).grid(row=1, column=0, columnspan=3, sticky="w", padx=6, pady=(6, 4))
        ttk.Checkbutton(
            self.advanced_frame,
            text="Abrir automaticamente a planilha exportada",
            variable=self.vars["auto_open_output"],
        ).grid(row=1, column=3, columnspan=3, sticky="w", padx=6, pady=(6, 4))

        self.expert_frame = ttk.LabelFrame(config_parent, text="Ajustes finos do algoritmo", padding=8)
        advanced_weight_fields = [
            ("Peso token-set", "weight_token_set", "Quanto a sobreposição de tokens influencia a pontuação final."),
            ("Peso parcial", "weight_partial", "Quanto a similaridade parcial de substring influencia a pontuação final."),
            ("Peso sort", "weight_sort", "Quanto a similaridade sem considerar ordem dos tokens influencia a pontuação final."),
            ("Peso prefixo", "weight_prefix", "Quanto o prefixo configurado contribui para a pontuação final."),
            ("Peso ordenado", "weight_ordered_chars", "Quanto a similaridade de caracteres em ordem contribui para a pontuação final."),
            ("Peso alinhado", "weight_aligned_chars", "Quanto a correspondência de caracteres na mesma posição contribui para a pontuação final."),
            ("Penalidade de tamanho", "length_gap_penalty_per_char", "Penalidade aplicada por caractere de diferença no tamanho do nome."),
            ("Penalidade máx. de tamanho", "max_length_gap_penalty", "Penalidade total máxima para grandes diferenças no tamanho do nome."),
            ("Penalidade de sobrenome", "missing_surname_penalty", "Penalidade quando o primeiro nome bate, mas a estrutura de sobrenome não."),
        ]
        for index, (label, key, tooltip) in enumerate(advanced_weight_fields, start=1):
            row = index // 2 + 2
            col = ((index - 1) % 2) * 3
            self._add_setting_field(self.expert_frame, row, col, label, key, tooltip)
        for column in (1, 4):
            self.expert_frame.columnconfigure(column, weight=1)

        button_row = ttk.Frame(config_parent)
        button_row.pack(fill="x", pady=(0, 6))
        ttk.Button(button_row, text="Preencher nome de saída", command=self.autofill_output_name, style="Info.TButton").pack(side="left")
        ttk.Button(button_row, text="Validar + Pré-visualizar planilha", command=self.validate_and_preview, style="Warning.TButton").pack(side="left", padx=8)
        ttk.Button(button_row, text="Salvar configuração visual", command=lambda: self.save_ui_state(show_feedback=True), style="Primary.TButton").pack(side="left", padx=8)
        ttk.Button(button_row, text="Iniciar análise", command=self.start_analysis, style="Success.TButton").pack(side="left", padx=8)

        preview_frame = ttk.LabelFrame(config_parent, text="Prévia da validação", padding=8)
        preview_frame.pack(fill="both", expand=True)
        self.validation_text = tk.Text(
            preview_frame,
            wrap="word",
            height=22,
            bg=UI_COLORS["field"],
            fg=UI_COLORS["text"],
            insertbackground=UI_COLORS["text"],
            relief="flat",
        )
        self.validation_text.pack(fill="both", expand=True)

    def _build_analyze_tab(self) -> None:
        actions = ttk.Frame(self.tab_analyze)
        actions.pack(fill="x", pady=(0, 8))
        self.analyze_button = ttk.Button(actions, text="Executar análise", command=self.start_analysis, style="Primary.TButton")
        self.analyze_button.pack(side="left")
        ttk.Button(actions, text="Atualizar prévia", command=self.refresh_analysis_views, style="Info.TButton").pack(side="left", padx=8)
        ttk.Label(actions, text="Visão rápida: cards + gráfico + tabela destacada por status.").pack(side="right")

        content = ttk.Panedwindow(self.tab_analyze, orient="vertical")
        content.pack(fill="both", expand=True)

        summary_frame = ttk.LabelFrame(content, text="Resumo", padding=8)
        preview_frame = ttk.LabelFrame(content, text="Prévia de resultados", padding=8)
        log_frame = ttk.LabelFrame(content, text="Log da análise", padding=8)
        content.add(summary_frame, weight=1)
        content.add(preview_frame, weight=3)
        content.add(log_frame, weight=2)

        cards = [
            ("Total de Linhas", "total"),
            ("Aceitos", "accepted"),
            ("Revisão", "review"),
            ("Sem Match", "no_match"),
        ]
        for index, (label, key) in enumerate(cards):
            card = ttk.LabelFrame(summary_frame, text=label, padding=10)
            card.grid(row=0, column=index, sticky="nsew", padx=6, pady=4)
            ttk.Label(card, textvariable=self.summary_card_vars[key], font=("Segoe UI", 18, "bold")).pack(anchor="center")
            summary_frame.columnconfigure(index, weight=1)
        self.analysis_chart_canvas = tk.Canvas(
            summary_frame,
            height=116,
            highlightthickness=0,
            bg=UI_COLORS["panel_alt"],
        )
        self.analysis_chart_canvas.grid(row=1, column=0, columnspan=4, sticky="ew", padx=6, pady=(8, 0))

        preview_columns = ("excel_row", "name", "analysis", "final", "match", "score", "flags")
        self.preview_tree = ttk.Treeview(preview_frame, columns=preview_columns, show="headings", height=14)
        headers = {
            "excel_row": "Linha",
            "name": "Nome T1",
            "analysis": "Análise",
            "final": "Final",
            "match": "T2 Sugerido",
            "score": "Pontuação",
            "flags": "Sinalizadores",
        }
        widths = {"excel_row": 70, "name": 240, "analysis": 100, "final": 100, "match": 240, "score": 80, "flags": 260}
        for column in preview_columns:
            self.preview_tree.heading(column, text=headers[column])
            self.preview_tree.column(column, width=widths[column], anchor="w")
        self._configure_status_tree(self.preview_tree)
        self.preview_tree.pack(fill="both", expand=True)

        self.log_text = tk.Text(
            log_frame,
            wrap="word",
            height=10,
            bg=UI_COLORS["field"],
            fg=UI_COLORS["text"],
            insertbackground=UI_COLORS["text"],
            relief="flat",
        )
        self.log_text.pack(fill="both", expand=True)

    def _build_review_tab(self) -> None:
        layout = ttk.Panedwindow(self.tab_review, orient="horizontal")
        layout.pack(fill="both", expand=True)

        queue_frame = ttk.LabelFrame(layout, text="Linhas que precisam de revisão", padding=8)
        detail_frame = ttk.LabelFrame(layout, text="Detalhes da linha selecionada", padding=8)
        layout.add(queue_frame, weight=2)
        layout.add(detail_frame, weight=3)

        queue_toolbar = ttk.Frame(queue_frame)
        queue_toolbar.pack(fill="x", pady=(0, 8))
        ttk.Label(queue_toolbar, text="Filtro").pack(side="left")
        review_filter = ttk.Combobox(
            queue_toolbar,
            textvariable=self.review_filter_var,
            values=["Todos", "Com conflito", "Gap baixo", "Realocado global", "Tamanho/posição", "Busca textual"],
            state="readonly",
            width=18,
        )
        review_filter.pack(side="left", padx=(8, 8))
        review_filter.bind("<<ComboboxSelected>>", lambda _event: self.refresh_review_views())
        ttk.Label(queue_toolbar, text="Busca").pack(side="left")
        review_search_entry = ttk.Entry(queue_toolbar, textvariable=self.review_search_var, width=28)
        review_search_entry.pack(side="left", padx=(8, 8))
        review_search_entry.bind("<KeyRelease>", lambda _event: self.refresh_review_views())
        ttk.Button(queue_toolbar, text="Limpar", command=self.clear_review_filters, style="Info.TButton").pack(side="left")

        queue_columns = ("row", "name", "status", "flags", "suggested")
        queue_tree_container = ttk.Frame(queue_frame)
        queue_tree_container.pack(fill="both", expand=True)
        self.review_tree = ttk.Treeview(queue_tree_container, columns=queue_columns, show="headings", height=24)
        for column, title, width in (
            ("row", "Linha", 70),
            ("name", "Nome T1", 220),
            ("status", "Status Final", 110),
            ("flags", "Sinalizadores", 230),
            ("suggested", "T2 Sugerido", 220),
        ):
            self.review_tree.heading(column, text=title)
            self.review_tree.column(column, width=width, anchor="w")
        review_tree_scrollbar = ttk.Scrollbar(queue_tree_container, orient="vertical", command=self.review_tree.yview)
        self.review_tree.configure(yscrollcommand=review_tree_scrollbar.set)
        self._configure_status_tree(self.review_tree)
        self.review_tree.pack(side="left", fill="both", expand=True)
        review_tree_scrollbar.pack(side="right", fill="y")
        self.review_tree.bind("<<TreeviewSelect>>", self.on_review_row_selected)

        detail_canvas = tk.Canvas(detail_frame, highlightthickness=0, bg=UI_COLORS["panel"])
        detail_scrollbar = ttk.Scrollbar(detail_frame, orient="vertical", command=detail_canvas.yview)
        detail_canvas.configure(yscrollcommand=detail_scrollbar.set)
        detail_scrollbar.pack(side="right", fill="y")
        detail_canvas.pack(side="left", fill="both", expand=True)
        detail_content = ttk.Frame(detail_canvas)
        detail_window_id = detail_canvas.create_window((0, 0), window=detail_content, anchor="nw")
        detail_content.bind(
            "<Configure>",
            lambda _event: detail_canvas.configure(scrollregion=detail_canvas.bbox("all")),
        )
        detail_canvas.bind(
            "<Configure>",
            lambda event: detail_canvas.itemconfigure(detail_window_id, width=event.width),
        )
        detail_canvas.bind("<Enter>", lambda _event: detail_canvas.bind_all("<MouseWheel>", lambda e: detail_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")))
        detail_canvas.bind("<Leave>", lambda _event: detail_canvas.unbind_all("<MouseWheel>"))

        ttk.Label(detail_content, textvariable=self.review_hint_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
        ttk.Label(detail_content, text="Prévia da linha").pack(anchor="w")
        self.review_detail_text = tk.Text(
            detail_content,
            height=10,
            wrap="word",
            bg=UI_COLORS["field"],
            fg=UI_COLORS["text"],
            insertbackground=UI_COLORS["text"],
            relief="flat",
        )
        self.review_detail_text.pack(fill="x", pady=(0, 8))

        ttk.Label(detail_content, text="Prévia de candidatos").pack(anchor="w")
        candidate_columns = ("rank", "candidate", "score", "ordered", "aligned", "auto", "review", "quota")
        candidate_container = ttk.Frame(detail_content)
        candidate_container.pack(fill="both", expand=True, pady=(0, 8))
        self.candidate_tree = ttk.Treeview(candidate_container, columns=candidate_columns, show="headings", height=14)
        for column, title, width in (
            ("rank", "Rank", 60),
            ("candidate", "Candidato T2", 260),
            ("score", "Pontuação", 80),
            ("ordered", "Ordenado", 80),
            ("aligned", "Alinhado", 80),
            ("auto", "Auto", 70),
            ("review", "Revisão", 70),
            ("quota", "Cota", 90),
        ):
            self.candidate_tree.heading(column, text=title)
            self.candidate_tree.column(column, width=width, anchor="w")
        candidate_tree_scrollbar = ttk.Scrollbar(candidate_container, orient="vertical", command=self.candidate_tree.yview)
        self.candidate_tree.configure(yscrollcommand=candidate_tree_scrollbar.set)
        self.candidate_tree.pack(side="left", fill="both", expand=True)
        candidate_tree_scrollbar.pack(side="right", fill="y")

        note_frame = ttk.Frame(detail_content)
        note_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(note_frame, text="Observação manual").pack(side="left")
        self.manual_note_entry = ttk.Entry(note_frame, textvariable=self.manual_note_var)
        self.manual_note_entry.pack(side="left", fill="x", expand=True, padx=(8, 0))

        buttons = ttk.Frame(detail_content)
        buttons.pack(fill="x")
        self.review_action_buttons: list[ttk.Button] = []
        accept_button = ttk.Button(buttons, text="Aceitar candidato selecionado", command=self.accept_selected_candidate, style="Success.TButton")
        accept_button.pack(side="left")
        self.review_action_buttons.append(accept_button)
        no_match_button = ttk.Button(buttons, text="Marcar sem match", command=self.mark_selected_no_match, style="Danger.TButton")
        no_match_button.pack(side="left", padx=8)
        self.review_action_buttons.append(no_match_button)
        keep_review_button = ttk.Button(buttons, text="Manter em revisão", command=self.keep_selected_in_review, style="Warning.TButton")
        keep_review_button.pack(side="left")
        self.review_action_buttons.append(keep_review_button)
        reset_button = ttk.Button(buttons, text="Resetar decisão manual", command=self.reset_manual_decision, style="Info.TButton")
        reset_button.pack(side="left", padx=8)
        self.review_action_buttons.append(reset_button)

    def _build_export_tab(self) -> None:
        actions = ttk.Frame(self.tab_export)
        actions.pack(fill="x", pady=(0, 8))
        self.export_button = ttk.Button(actions, text="Exportar resultados revisados", command=self.start_export, style="Success.TButton")
        self.export_button.pack(side="left")
        ttk.Button(actions, text="Abrir última exportação", command=self.open_last_output, style="Primary.TButton").pack(side="left", padx=8)
        ttk.Label(actions, textvariable=self.export_snapshot_var).pack(side="right")

        info_frame = ttk.LabelFrame(self.tab_export, text="Status da exportação", padding=8)
        info_frame.pack(fill="both", expand=True)
        self.export_chart_canvas = tk.Canvas(
            info_frame,
            height=116,
            highlightthickness=0,
            bg=UI_COLORS["panel_alt"],
        )
        self.export_chart_canvas.pack(fill="x", pady=(0, 8))
        self.export_text = tk.Text(
            info_frame,
            wrap="word",
            height=28,
            bg=UI_COLORS["field"],
            fg=UI_COLORS["text"],
            insertbackground=UI_COLORS["text"],
            relief="flat",
        )
        self.export_text.pack(fill="both", expand=True)

    def _add_file_row(self, parent: ttk.LabelFrame, label: str, var_key: str, command: Callable[[], None], row: int) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=4, pady=3)
        ttk.Entry(parent, textvariable=self.vars[var_key], width=88).grid(row=row, column=1, sticky="ew", padx=4, pady=3)
        ttk.Button(parent, text="Procurar...", command=command, style="Info.TButton").grid(row=row, column=2, padx=4, pady=3)
        parent.columnconfigure(1, weight=1)

    def _add_sheet_field(
        self,
        parent: ttk.LabelFrame,
        row: int,
        col: int,
        label: str,
        var_key: str,
        tooltip: str,
    ) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=6, pady=5)
        combobox = ttk.Combobox(parent, textvariable=self.vars[var_key], state="readonly", width=22)
        combobox.grid(row=row, column=col + 1, sticky="ew", padx=4, pady=3)
        combobox.bind("<<ComboboxSelected>>", lambda _event: self.refresh_sheet_choices())
        if var_key == "sheet_t1":
            self.sheet_t1_combo = combobox
        else:
            self.sheet_t2_combo = combobox
        info = ttk.Label(parent, text="?", width=2, anchor="center")
        info.grid(row=row, column=col + 2, sticky="w", padx=(0, 6), pady=3)
        ToolTip(info, tooltip)

    def _add_combobox_field(
        self,
        parent: ttk.LabelFrame,
        row: int,
        col: int,
        label: str,
        var_key: str,
        values: list[str],
        tooltip: str,
    ) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=4, pady=3)
        ttk.Combobox(parent, textvariable=self.vars[var_key], values=values, state="readonly", width=24).grid(
            row=row,
            column=col + 1,
            sticky="ew",
            padx=4,
            pady=3,
        )
        info = ttk.Label(parent, text="?", width=2, anchor="center")
        info.grid(row=row, column=col + 2, sticky="w", padx=(0, 6), pady=3)
        ToolTip(info, tooltip)

    def _add_color_field(self, parent: ttk.LabelFrame, row: int, col: int, label: str, var_key: str) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=4, pady=3)
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=col + 1, sticky="ew", padx=4, pady=3)
        combo = ttk.Combobox(
            frame,
            textvariable=self.vars[var_key],
            values=MATCH_COLOR_LABEL_OPTIONS,
            width=24,
        )
        combo.pack(side="left", fill="x", expand=True)
        combo.bind("<<ComboboxSelected>>", lambda _event, key=var_key: self.normalize_color_var_value(key, save_state=True))
        combo.bind("<FocusOut>", lambda _event, key=var_key: self.normalize_color_var_value(key))
        ttk.Button(
            frame,
            text="Cor...",
            command=lambda key=var_key: self.choose_color(key),
            style="Primary.TButton",
        ).pack(side="left", padx=(6, 0))

    def normalize_color_var_value(self, var_key: str, save_state: bool = False) -> None:
        current = str(self.vars[var_key].get() or "").strip()
        formatted = format_color_for_ui(current, DEFAULT_MATCH_COLORS[var_key])
        if formatted != current:
            self.vars[var_key].set(formatted)
        if save_state:
            self.save_ui_state()

    def _add_csv_column_field(
        self,
        parent: ttk.LabelFrame,
        row: int,
        col: int,
        label: str,
        var_key: str,
        tooltip: str,
        side: str,
    ) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=4, pady=3)
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=col + 1, sticky="ew", padx=4, pady=3)
        ttk.Entry(frame, textvariable=self.vars[var_key], width=24).pack(side="left", fill="x", expand=True)
        ttk.Button(
            frame,
            text="Sugerir",
            style="Info.TButton",
            command=lambda key=var_key, side_key=side: self.apply_tab4_default_columns(key, side_key),
        ).pack(side="left", padx=(6, 0))
        ttk.Button(
            frame,
            text="Todas",
            style="Primary.TButton",
            command=lambda key=var_key, side_key=side: self.apply_tab4_all_columns(key, side_key),
        ).pack(side="left", padx=(4, 0))
        info = ttk.Label(parent, text="?", width=2, anchor="center")
        info.grid(row=row, column=col + 2, sticky="w", padx=(0, 6), pady=3)
        ToolTip(info, tooltip)

    def choose_color(self, var_key: str) -> None:
        current = str(self.vars[var_key].get() or "")
        color = f"#{normalize_fill_color(current, DEFAULT_MATCH_COLORS[var_key])[-6:]}"
        chosen = colorchooser.askcolor(color=color, title=f"Escolher cor para {var_key}")
        if chosen and chosen[1]:
            self.vars[var_key].set(format_color_for_ui(chosen[1], DEFAULT_MATCH_COLORS[var_key]))
            self.save_ui_state()

    def apply_tab4_default_columns(self, var_key: str, side: str) -> None:
        options = self.column_options_t1 if side == "t1" else self.column_options_t2
        match_key = "match2_col_t1" if side == "t1" else "match2_col_t2"
        selected = parse_csv_columns(self.vars[var_key].get())
        if self.vars[match_key].get():
            idx = excel_col_to_index(self.vars[match_key].get())
            if idx < len(options):
                default_col = options[idx]
                if default_col not in selected:
                    selected.insert(0, default_col)
        serialized = serialize_csv_columns(selected)
        if serialized != self.vars[var_key].get():
            self.vars[var_key].set(serialized)
            self.save_ui_state()

    def apply_tab4_all_columns(self, var_key: str, side: str) -> None:
        options = self.column_options_t1 if side == "t1" else self.column_options_t2
        self.vars[var_key].set(serialize_csv_columns(options))
        self.apply_tab4_default_columns(var_key, side)

    def refresh_sheet_choices(self) -> None:
        metadata_errors: list[str] = []
        self.sheet_options_t1 = list_workbook_sheets(self.vars["input_file_t1"].get(), errors=metadata_errors)
        self.sheet_options_t2 = list_workbook_sheets(self.vars["input_file_t2"].get(), errors=metadata_errors)
        if hasattr(self, "sheet_t1_combo"):
            self.sheet_t1_combo["values"] = self.sheet_options_t1
        if hasattr(self, "sheet_t2_combo"):
            self.sheet_t2_combo["values"] = self.sheet_options_t2
        if self.sheet_options_t1 and self.vars["sheet_t1"].get() not in self.sheet_options_t1:
            self.vars["sheet_t1"].set(self.sheet_options_t1[0])
        if self.sheet_options_t2 and self.vars["sheet_t2"].get() not in self.sheet_options_t2:
            self.vars["sheet_t2"].set(self.sheet_options_t2[0])
        if not self.sheet_options_t1:
            self.vars["sheet_t1"].set("")
        if not self.sheet_options_t2:
            self.vars["sheet_t2"].set("")
        try:
            header_t1 = int(self.vars["header_row_t1"].get() or 1)
        except ValueError:
            header_t1 = 1
        try:
            header_t2 = int(self.vars["header_row_t2"].get() or 1)
        except ValueError:
            header_t2 = 1
        self.column_options_t1 = list_workbook_columns(
            self.vars["input_file_t1"].get(),
            self.vars["sheet_t1"].get(),
            header_t1,
            errors=metadata_errors,
        )
        self.column_options_t2 = list_workbook_columns(
            self.vars["input_file_t2"].get(),
            self.vars["sheet_t2"].get(),
            header_t2,
            errors=metadata_errors,
        )
        if metadata_errors:
            signature = " | ".join(metadata_errors)
            if signature != self._last_workbook_metadata_error:
                self.log("Falha ao carregar metadados das planilhas:")
                for message in metadata_errors:
                    self.log(f"- {message}")
                self.set_status("Falha ao ler metadados da planilha. Verifique o log para detalhes.", None)
                self._last_workbook_metadata_error = signature
        else:
            self._last_workbook_metadata_error = ""
        self.apply_tab4_default_columns("tab4_extra_cols_t1", "t1")
        self.apply_tab4_default_columns("tab4_extra_cols_t2", "t2")

    def _add_setting_field(
        self,
        parent,
        row: int,
        col: int,
        label: str,
        var_key: str,
        tooltip: str,
    ) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=4, pady=3)
        entry = ttk.Entry(parent, textvariable=self.vars[var_key], width=22)
        entry.grid(row=row, column=col + 1, sticky="ew", padx=4, pady=3)
        if var_key in {"header_row_t1", "header_row_t2", "match2_col_t1", "match2_col_t2"}:
            entry.bind("<FocusOut>", lambda _event: self.refresh_sheet_choices())
        info = ttk.Label(parent, text="?", width=2, anchor="center")
        info.grid(row=row, column=col + 2, sticky="w", padx=(0, 6), pady=3)
        ToolTip(info, tooltip)

    def clear_review_filters(self) -> None:
        self.review_filter_var.set("Todos")
        self.review_search_var.set("")
        self.refresh_review_views()

    def log(self, message: str) -> None:
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def set_status(self, message: str, percent: float | None = None) -> None:
        self.status_var.set(message)
        if percent is not None:
            self.progress_var.set(percent)

    def pick_input_file_t1(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecionar planilha Excel 1",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")],
        )
        if path:
            self.vars["input_file_t1"].set(path)
            self.refresh_sheet_choices()
            self.autofill_output_name()
            self.save_ui_state()
            self.log(f"Planilha Excel 1 selecionada: {path}")

    def pick_input_file_t2(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecionar planilha Excel 2",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")],
        )
        if path:
            self.vars["input_file_t2"].set(path)
            self.refresh_sheet_choices()
            self.autofill_output_name()
            self.save_ui_state()
            self.log(f"Planilha Excel 2 selecionada: {path}")

    def pick_output_file(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Salvar planilha de saída como",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=Path(self.vars["output_file"].get() or "resultado_matching.xlsx").name,
        )
        if path:
            self.vars["output_file"].set(path)
            self.save_ui_state()
            self.log(f"Planilha de saída definida para: {path}")

    def autofill_output_name(self) -> None:
        input_t1 = self.vars["input_file_t1"].get().strip()
        input_t2 = self.vars["input_file_t2"].get().strip()
        if not input_t1 and not input_t2:
            self.vars["output_file"].set("resultado_matching.xlsx")
            return
        left = Path(input_t1).stem if input_t1 else "excel1"
        right = Path(input_t2).stem if input_t2 else "excel2"
        base_path = Path(input_t1 or input_t2)
        self.vars["output_file"].set(str(base_path.with_name(f"{left}__{right}_resultado_matching.xlsx")))

    def collect_config_from_vars(self) -> dict[str, Any]:
        return {
            key: variable.get().strip() if isinstance(variable, tk.StringVar) else variable.get()
            for key, variable in self.vars.items()
        }

    def validate_and_preview(self) -> None:
        try:
            config = validate_config(self.collect_config_from_vars(), validate_workbook=True)
            preview = collect_workbook_preview(config)
            self.validation_text.delete("1.0", "end")
            self.validation_text.insert("1.0", preview)
            self.set_status("Validação da planilha concluída com sucesso.", 0)
            self.export_snapshot_var.set("Mapeamento validado e pronto para análise.")
            self.save_ui_state()
            self.log("Prévia da validação da planilha atualizada.")
            self.notebook.select(self.tab_config)
        except Exception as exc:
            self._show_error("Erro de validação", str(exc))

    def start_analysis(self) -> None:
        try:
            config = validate_config(self.collect_config_from_vars(), validate_workbook=True)
        except Exception as exc:
            self._show_error("Erro de validação", str(exc))
            return

        self.set_busy(True, operation="analysis")
        self.set_status("Iniciando análise...", 0)
        self.log("Iniciando análise.")

        def worker() -> None:
            try:
                result = analyze_matching(config, progress_callback=self.safe_progress)
                self.root.after(0, lambda: self.on_analysis_success(result))
            except Exception as exc:
                tb = traceback.format_exc()
                self.root.after(0, self.on_background_error, str(exc), tb)

        threading.Thread(target=worker, daemon=True).start()

    def safe_progress(self, message: str, percent: float | None = None) -> None:
        self.root.after(0, lambda: self._update_progress(message, percent))

    def _update_progress(self, message: str, percent: float | None = None) -> None:
        self.set_status(message, percent)
        self.log(message)

    def set_busy(self, is_busy: bool, operation: str | None = None) -> None:
        if is_busy:
            self.busy_operation = operation or "background"
        else:
            self.busy_operation = None
        state = "disabled" if is_busy else "normal"
        self.analyze_button.config(state=state)
        self.export_button.config(state=state)
        review_controls_state = "disabled" if is_busy and self.busy_operation == "export" else "normal"
        for button in getattr(self, "review_action_buttons", []):
            button.config(state=review_controls_state)
        if hasattr(self, "manual_note_entry"):
            self.manual_note_entry.config(state=review_controls_state)

    def _ensure_review_mutation_allowed(self) -> bool:
        if self.busy_operation == "export":
            self._show_warning("Exportação em andamento", "Aguarde o fim da exportação para alterar decisões manuais.")
            return False
        return True

    def on_analysis_success(self, result: AnalysisResult) -> None:
        self.analysis_result = result
        self.catalog_df = result.catalog_df.copy()
        self.last_output_path = None
        self.set_busy(False)
        self.set_status("Análise concluída.", 100)
        self.refresh_analysis_views()
        self.refresh_review_views()
        self.refresh_export_view()
        self.notebook.select(self.tab_analyze)
        self.log("Análise concluída com sucesso.")

    def on_background_error(self, error_message: str, tb: str) -> None:
        self.set_busy(False)
        self.set_status("Operação falhou.", 0)
        self.log("ERRO:")
        self.log(error_message)
        self.log(tb)
        self._show_error("Erro", error_message)

    def refresh_analysis_views(self) -> None:
        if self.analysis_result is None:
            return

        summary_df = build_summary(self.analysis_result.results_df)
        self.analysis_result.summary_df = summary_df
        total_rows = len(self.analysis_result.results_df)
        counts = summary_df.set_index("status")["quantidade"].to_dict() if not summary_df.empty else {}
        self.summary_card_vars["total"].set(str(total_rows))
        self.summary_card_vars["accepted"].set(str(int(counts.get("ACEITO", 0))))
        self.summary_card_vars["review"].set(str(int(counts.get("REVISAR", 0))))
        self.summary_card_vars["no_match"].set(str(int(counts.get("SEM_MATCH", 0))))
        self._draw_summary_chart(self.analysis_chart_canvas, counts)

        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        preview_df = self.analysis_result.results_df.head(50)
        for row in preview_df.itertuples(index=False):
            tag = getattr(row, "final_status", "") or getattr(row, "analysis_status", "")
            if getattr(row, "final_conflict_flags", ""):
                tag = "CONFLITO"
            self.preview_tree.insert(
                "",
                "end",
                values=(
                    getattr(row, "excel_row_t1", ""),
                    getattr(row, "nome_t1_original", ""),
                    getattr(row, "analysis_status", ""),
                    getattr(row, "final_status", ""),
                    getattr(row, "final_match_t2_original", ""),
                    getattr(row, "final_score", ""),
                    getattr(row, "final_conflict_flags", ""),
                ),
                tags=(tag,),
            )

    def refresh_review_views(self) -> None:
        if self.analysis_result is None:
            return

        recompute_final_state(self.analysis_result.results_df, self.catalog_df, config=self.analysis_result.config)
        review_df = self.analysis_result.results_df[self.analysis_result.results_df["final_status"] == "REVISAR"].copy()
        filter_value = self.review_filter_var.get()
        search_value = normalize_name(self.review_search_var.get())
        if filter_value == "Com conflito":
            review_df = review_df[review_df["final_conflict_flags"].fillna("") != ""]
        elif filter_value == "Gap baixo":
            review_df = review_df[review_df["final_conflict_flags"].fillna("").str.contains("LOW_GAP", na=False)]
        elif filter_value == "Realocado global":
            review_df = review_df[review_df["final_conflict_flags"].fillna("").str.contains("GLOBAL_REALLOCATED", na=False)]
        elif filter_value == "Tamanho/posição":
            review_df = review_df[review_df["final_conflict_flags"].fillna("").str.contains("LENGTH_POSITION_REVIEW", na=False)]
        if search_value:
            review_df = review_df[
                review_df["nome_t1_original"].fillna("").apply(normalize_name).str.contains(search_value, na=False)
                | review_df["final_match_t2_original"].fillna("").apply(normalize_name).str.contains(search_value, na=False)
            ]
        self.analysis_result.review_df = review_df.copy()

        for item in self.review_tree.get_children():
            self.review_tree.delete(item)
        for row in self.analysis_result.review_df.itertuples(index=False):
            tag = "CONFLITO" if getattr(row, "final_conflict_flags", "") else getattr(row, "final_status", "")
            self.review_tree.insert(
                "",
                "end",
                iid=str(getattr(row, "source_row_id")),
                values=(
                    getattr(row, "excel_row_t1", ""),
                    getattr(row, "nome_t1_original", ""),
                    getattr(row, "final_status", ""),
                    getattr(row, "final_conflict_flags", ""),
                    getattr(row, "final_match_t2_original", ""),
                ),
                tags=(tag,),
            )

        self.review_detail_text.delete("1.0", "end")
        self.review_hint_var.set(
            f"Pendências exibidas: {len(self.analysis_result.review_df)} • filtro atual: {self.review_filter_var.get()}"
        )
        for item in self.candidate_tree.get_children():
            self.candidate_tree.delete(item)

    def on_review_row_selected(self, _event=None) -> None:
        if self.analysis_result is None:
            return
        selection = self.review_tree.selection()
        if not selection:
            return

        source_row_id = int(selection[0])
        row_df = self.analysis_result.results_df[self.analysis_result.results_df["source_row_id"] == source_row_id]
        if row_df.empty:
            return
        row = row_df.iloc[0]
        self.manual_note_var.set(str(row.get("manual_note", "") or ""))
        self.review_hint_var.set(
            f"Linha {row['excel_row_t1']} • status {row['final_status']} • sinalizadores: {row['final_conflict_flags'] or 'nenhum'}"
        )

        details = [
            f"Linha no Excel: {row['excel_row_t1']}",
            f"T1 original: {row['nome_t1_original']}",
            f"T1 normalizado: {row['nome_t1_norm']}",
            f"Status da análise: {row['analysis_status']}",
            f"Método da análise: {row['analysis_method']}",
            f"T2 sugerido: {row['analysis_match_t2_original']}",
            f"Pontuação sugerida (nome): {row['analysis_score']}",
            f"Pontuação composta (nome + match2): {row.get('analysis_score_composite', '')}",
            f"Match 2 T1: {row.get('analysis_match2_t1', '')}",
            f"Match 2 T2: {row.get('analysis_match2_t2', '')}",
            f"Match 2 prefixo igual: {row.get('analysis_match2_equal_prefix', False)}",
            f"Match 2 score: {row.get('analysis_match2_score', '')}",
            f"Motivo da revisão: {row['analysis_review_reason']}",
            f"Sinalizadores de conflito: {row['final_conflict_flags']}",
        ]
        self.review_detail_text.delete("1.0", "end")
        self.review_detail_text.insert("1.0", "\n".join(details))

        for item in self.candidate_tree.get_children():
            self.candidate_tree.delete(item)

        candidates = self.analysis_result.candidates_df[
            self.analysis_result.candidates_df["source_row_id"] == source_row_id
        ].sort_values(["rank", "score_composite"], ascending=[True, False])
        for candidate in candidates.itertuples(index=False):
            quota_text = f"{getattr(candidate, 'quota_limit', '')}"
            candidate_tag = "ACEITO" if getattr(candidate, "rank") == 1 else ""
            self.candidate_tree.insert(
                "",
                "end",
                iid=f"{source_row_id}:{getattr(candidate, 'rank')}",
                values=(
                    getattr(candidate, "rank"),
                    getattr(candidate, "nome_t2_original"),
                    getattr(candidate, "score_composite"),
                    getattr(candidate, "score_ordered_chars"),
                    getattr(candidate, "score_aligned_chars"),
                    "S" if getattr(candidate, "eligible_for_global") else "",
                    "S" if getattr(candidate, "review_eligible") else "",
                    quota_text,
                ),
                tags=(candidate_tag,),
            )

    def _selected_source_row_id(self) -> int | None:
        selection = self.review_tree.selection()
        if not selection:
            self._show_warning("Revisão", "Selecione primeiro uma linha da fila de revisão.")
            return None
        return int(selection[0])

    def accept_selected_candidate(self) -> None:
        if self.analysis_result is None:
            return
        if not self._ensure_review_mutation_allowed():
            return
        source_row_id = self._selected_source_row_id()
        if source_row_id is None:
            return
        candidate_selection = self.candidate_tree.selection()
        if not candidate_selection:
            self._show_warning("Revisão", "Selecione um candidato para aceitar.")
            return

        rank = int(candidate_selection[0].split(":")[1])
        candidate_df = self.analysis_result.candidates_df[
            (self.analysis_result.candidates_df["source_row_id"] == source_row_id)
            & (self.analysis_result.candidates_df["rank"] == rank)
        ]
        if candidate_df.empty:
            return
        candidate = candidate_df.iloc[0]
        row_mask = self.analysis_result.results_df["source_row_id"] == source_row_id
        self.manual_sequence += 1
        self.analysis_result.results_df.loc[row_mask, "manual_status"] = "ACCEPT"
        self.analysis_result.results_df.loc[row_mask, "manual_match_t2_original"] = candidate["nome_t2_original"]
        self.analysis_result.results_df.loc[row_mask, "manual_match_t2_norm"] = candidate["nome_t2_norm"]
        self.analysis_result.results_df.loc[row_mask, "manual_line_match_t2"] = candidate["excel_row_t2"]
        self.analysis_result.results_df.loc[row_mask, "manual_score"] = candidate["score_composite"]
        self.analysis_result.results_df.loc[row_mask, "manual_note"] = self.manual_note_var.get().strip()
        self.analysis_result.results_df.loc[row_mask, "manual_sequence"] = self.manual_sequence
        recompute_final_state(self.analysis_result.results_df, self.catalog_df, config=self.analysis_result.config)
        self.last_manual_actions.insert(0, f"Linha {source_row_id}: aceite manual pelo candidato rank {rank}")
        self.last_manual_actions = self.last_manual_actions[:8]
        self.refresh_analysis_views()
        self.refresh_review_views()
        self.refresh_export_view()
        self.log(f"Aceite manual aplicado à linha {source_row_id} usando o candidato rank {rank}.")

    def mark_selected_no_match(self) -> None:
        if self.analysis_result is None:
            return
        if not self._ensure_review_mutation_allowed():
            return
        source_row_id = self._selected_source_row_id()
        if source_row_id is None:
            return
        row_mask = self.analysis_result.results_df["source_row_id"] == source_row_id
        self.manual_sequence += 1
        self.analysis_result.results_df.loc[row_mask, "manual_status"] = "NO_MATCH"
        self.analysis_result.results_df.loc[row_mask, "manual_match_t2_original"] = ""
        self.analysis_result.results_df.loc[row_mask, "manual_match_t2_norm"] = ""
        self.analysis_result.results_df.loc[row_mask, "manual_line_match_t2"] = pd.NA
        self.analysis_result.results_df.loc[row_mask, "manual_score"] = pd.NA
        self.analysis_result.results_df.loc[row_mask, "manual_note"] = self.manual_note_var.get().strip()
        self.analysis_result.results_df.loc[row_mask, "manual_sequence"] = self.manual_sequence
        recompute_final_state(self.analysis_result.results_df, self.catalog_df, config=self.analysis_result.config)
        self.last_manual_actions.insert(0, f"Linha {source_row_id}: marcada manualmente como sem match")
        self.last_manual_actions = self.last_manual_actions[:8]
        self.refresh_analysis_views()
        self.refresh_review_views()
        self.refresh_export_view()
        self.log(f"Linha {source_row_id} marcada manualmente como sem match.")

    def keep_selected_in_review(self) -> None:
        if self.analysis_result is None:
            return
        if not self._ensure_review_mutation_allowed():
            return
        source_row_id = self._selected_source_row_id()
        if source_row_id is None:
            return
        row_mask = self.analysis_result.results_df["source_row_id"] == source_row_id
        self.manual_sequence += 1
        self.analysis_result.results_df.loc[row_mask, "manual_status"] = "REVIEW"
        self.analysis_result.results_df.loc[row_mask, "manual_note"] = self.manual_note_var.get().strip()
        self.analysis_result.results_df.loc[row_mask, "manual_sequence"] = self.manual_sequence
        recompute_final_state(self.analysis_result.results_df, self.catalog_df, config=self.analysis_result.config)
        self.last_manual_actions.insert(0, f"Linha {source_row_id}: mantida em revisão")
        self.last_manual_actions = self.last_manual_actions[:8]
        self.refresh_analysis_views()
        self.refresh_review_views()
        self.refresh_export_view()
        self.log(f"Linha {source_row_id} mantida em revisão manual.")

    def reset_manual_decision(self) -> None:
        if self.analysis_result is None:
            return
        if not self._ensure_review_mutation_allowed():
            return
        source_row_id = self._selected_source_row_id()
        if source_row_id is None:
            return
        row_mask = self.analysis_result.results_df["source_row_id"] == source_row_id
        for column in [
            "manual_status",
            "manual_match_t2_original",
            "manual_match_t2_norm",
            "manual_note",
        ]:
            self.analysis_result.results_df.loc[row_mask, column] = ""
        self.analysis_result.results_df.loc[row_mask, "manual_line_match_t2"] = pd.NA
        self.analysis_result.results_df.loc[row_mask, "manual_score"] = pd.NA
        self.analysis_result.results_df.loc[row_mask, "manual_sequence"] = pd.NA
        recompute_final_state(self.analysis_result.results_df, self.catalog_df, config=self.analysis_result.config)
        self.last_manual_actions.insert(0, f"Linha {source_row_id}: decisão manual resetada")
        self.last_manual_actions = self.last_manual_actions[:8]
        self.refresh_analysis_views()
        self.refresh_review_views()
        self.refresh_export_view()
        self.log(f"Decisão manual resetada para a linha {source_row_id}.")

    def refresh_export_view(self) -> None:
        self.export_text.delete("1.0", "end")
        if self.analysis_result is None:
            self.export_text.insert("1.0", "Execute uma análise para preencher os detalhes da exportação.")
            return

        recompute_final_state(self.analysis_result.results_df, self.catalog_df, config=self.analysis_result.config)
        summary_df = build_summary(self.analysis_result.results_df)
        quota_df = build_quota_summary(self.analysis_result.results_df, self.catalog_df)
        unresolved = int((self.analysis_result.results_df["final_status"] == "REVISAR").sum())
        conflict_count = int((self.analysis_result.results_df["final_conflict_flags"].fillna("") != "").sum())
        counts = summary_df.set_index("status")["quantidade"].to_dict() if not summary_df.empty else {}
        self._draw_summary_chart(self.export_chart_canvas, counts)
        lines = ["Estado atual da exportação", "==========================", ""]
        for row in summary_df.itertuples(index=False):
            lines.append(f"{row.status}: {row.quantidade} ({row.percentual:.2f}%)")
        lines.extend(
            [
                "",
                f"Linhas ainda em revisão: {unresolved}",
                f"Linhas com sinalizadores de conflito: {conflict_count}",
                f"Planilha de saída: {self.vars['output_file'].get()}",
                "",
                "Últimas decisões manuais:",
            ]
        )
        if self.last_manual_actions:
            lines.extend(f"- {item}" for item in self.last_manual_actions[:5])
        else:
            lines.append("- Nenhuma decisão manual registrada nesta sessão.")
        lines.extend(
            [
                "",
                "Cotas mais preenchidas:",
            ]
        )
        if quota_df.empty:
            lines.append("- Nenhum catálogo de T2 disponível.")
        else:
            top_quota = quota_df.sort_values(["accepted_count", "nome_t2_original"], ascending=[False, True]).head(10)
            for row in top_quota.itertuples(index=False):
                lines.append(
                    f"- {row.nome_t2_original}: {row.accepted_count}/{row.quota_limit} aceitos"
                )
        self.export_text.insert("1.0", "\n".join(lines))

    def start_export(self) -> None:
        if self.analysis_result is None:
            self._show_warning("Exportação", "Execute a análise antes de exportar.")
            return
        try:
            config = validate_config(self.collect_config_from_vars(), validate_workbook=False)
            self.analysis_result.config["output_file"] = config["output_file"]
        except Exception as exc:
            self._show_error("Validação da exportação", str(exc))
            return

        self.set_busy(True, operation="export")
        self.set_status("Exportando planilha...", 0)
        self.log("Iniciando exportação.")

        def worker() -> None:
            try:
                output_path = export_analysis_result(
                    self.analysis_result,
                    output_file=self.analysis_result.config["output_file"],
                    progress_callback=self.safe_progress,
                )
                self.root.after(0, lambda: self.on_export_success(output_path))
            except Exception as exc:
                tb = traceback.format_exc()
                self.root.after(0, self.on_background_error, str(exc), tb)

        threading.Thread(target=worker, daemon=True).start()

    def on_export_success(self, output_path: Path) -> None:
        self.set_busy(False)
        self.last_output_path = output_path
        self.refresh_export_view()
        self.export_snapshot_var.set(f"Última exportação: {output_path.name}")
        self.save_ui_state()
        self.set_status(f"Exportação concluída: {output_path}", 100)
        self.log(f"Exportação concluída: {output_path}")
        if self.vars["auto_open_output"].get():
            try:
                open_file_with_default_app(output_path)
                self.log("Planilha exportada aberta automaticamente.")
            except Exception as exc:
                self.log(f"Não foi possível abrir automaticamente a planilha: {exc}")
        self._show_info("Exportação concluída", f"Planilha gerada:\n\n{output_path}")

    def open_last_output(self) -> None:
        if not self.last_output_path:
            self._show_warning("Abrir exportação", "Ainda não há planilha exportada disponível.")
            return
        try:
            open_file_with_default_app(self.last_output_path)
        except Exception as exc:
            self._show_error("Abrir exportação", str(exc))


def main() -> None:
    ensure_supported_python_version()
    root = ttk.Window(themename="darkly") if HAS_TTKBOOTSTRAP else tk.Tk()
    try:
        root.iconbitmap(default="")
    except Exception:
        pass
    app = MatcherApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
